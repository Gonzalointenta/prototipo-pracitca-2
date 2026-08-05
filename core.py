# -*- coding: utf-8 -*-
"""
core.py
Lógica de negocio para el sistema de registro de bodega municipal.
Implementa el diseño del documento "Diagnóstico inicial de área de inventario":
  - catálogo maestro + tabla de alias asociativos (búsqueda por lenguaje libre)
  - solicitud en 2 tiempos: preliminar (antes de firma física) -> editable -> cerrada
  - validación de disponibilidad (insumo=0 / insumo<solicitado / insumo>=solicitado)
  - alertas de stock agotado / bajo stock crítico
"""

import hashlib
import os
import re
import sqlite3
import time
import unicodedata
import uuid
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
from rapidfuzz import fuzz, process

import supabase_pg

DB_PATH = "bodega.db"

# Base compartida en la nube (Supabase/Postgres, región São Paulo): si esta
# variable está definida, get_connection() la usa SIEMPRE, ignorando el
# db_path que le pasen — así la web y la app de escritorio leen y escriben la
# misma base, en vez de cada una tener su propio archivo SQLite aislado. Sin
# ella (dev local, tests) se sigue usando el archivo SQLite de siempre.
def _leer_supabase_url():
    """
    URL de conexión a la base compartida en Supabase.

    En Streamlit Cloud los secrets se leen de forma confiable con st.secrets.
    También se exponen como variables de entorno, PERO esa exposición es
    perezosa: puede no estar poblada todavía en el momento en que se importa
    este módulo. Por eso consultar os.environ directo a veces devolvía None en
    la web —aunque el secret estuviera bien puesto— y la app caía a una base
    SQLite local vacía, sin la nómina de correos, dando "correo no autorizado"
    para correos que sí estaban cargados. Se consulta st.secrets primero (que
    fuerza la carga y es el camino confiable en la nube) y se cae a os.environ
    para la app de escritorio (el .exe), scripts y tests, que no corren dentro
    de Streamlit y usan la variable de entorno de siempre.

    Es el mismo patrón que ya usaba _password_encargado() para la contraseña
    del encargado por exactamente esta razón.
    """
    try:
        import streamlit as st
        if "SUPABASE_DB_URL" in st.secrets:
            return str(st.secrets["SUPABASE_DB_URL"])
    except Exception:
        # fuera de Streamlit (o sin secrets configurados) no hay st.secrets:
        # se sigue al entorno
        pass
    return os.environ.get("SUPABASE_DB_URL")


SUPABASE_DB_URL = _leer_supabase_url()

TZ_CHILE = ZoneInfo("America/Santiago")


def ahora_chile() -> datetime:
    """
    Hora actual en horario de Chile continental. Se usa en vez de
    datetime.now() en todo el sistema porque el servidor donde corre la app
    (Streamlit Cloud) no está en horario chileno. Usa la base de husos
    horarios IANA (paquete 'tzdata'), así que el cambio de horario de
    verano/invierno se resuelve solo, sin fechas hardcodeadas.
    """
    return datetime.now(TZ_CHILE).replace(tzinfo=None)


def hoy_chile():
    """Fecha (sin hora) de hoy en Chile."""
    return ahora_chile().date()


# ---------------------------------------------------------------- utilidades

def normalizar(texto: str) -> str:
    """Minúsculas, sin tildes, sin espacios sobrantes."""
    texto = texto.lower().strip()
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    return texto


def valor_unitario_desde(valor_saldo, saldo) -> float:
    """
    Valor unitario (en pesos) de un producto = valor total del saldo dividido
    por la cantidad, calculado UNA sola vez al momento del escaneo/corte.

    SMC solo informa el valor TOTAL del saldo por código (no un precio
    unitario), así que este cociente es la única forma de tener un valor por
    unidad. Se fija al cargar el catálogo y NO se recalcula al hacer
    solicitudes: a medida que se entrega stock se descuenta este valor fijo
    del total (ver cerrar_solicitud), de modo que el inventario en pesos baja
    solo y cada movimiento tiene un valor en plata consistente con el corte.
    Recalcularlo con el saldo ya movido daría un precio distinto y produciría
    disconformidades con SMC — por eso se calcula solo acá, con los valores
    del corte.

    Si el saldo del corte es 0 (o falta el valor), no hay de dónde derivar un
    unitario y se devuelve 0.
    """
    try:
        saldo = float(saldo)
        valor_saldo = float(valor_saldo)
    except (TypeError, ValueError):
        return 0.0
    if saldo <= 0 or valor_saldo <= 0:
        return 0.0
    return valor_saldo / saldo


def _normalizar_columna_mensaje(df: pd.DataFrame) -> pd.DataFrame:
    """
    Deja la columna 'mensaje_sistema' con "" (cadena vacía) donde no hay
    observación, en vez del NaN que devuelve pandas al leer un NULL.

    Por qué importa: los productos que tienen stock suficiente se guardan con
    mensaje_sistema en NULL. Al leerlos, pandas los entrega como NaN, y como
    en Python bool(nan) es True, el código de la interfaz que hacía
    `if fila["mensaje_sistema"]:` los tomaba como si tuvieran mensaje y
    mostraba literalmente el texto "nan" al lado de cada producto sano —tanto
    en la vista del solicitante como en la del encargado—. Con "" (que es
    falsy) esas comprobaciones vuelven a funcionar y las celdas quedan en
    blanco. Se aplica en el origen (las funciones que arman estas tablas) para
    que todas las vistas queden bien sin repetir el arreglo en cada una.
    """
    if "mensaje_sistema" in df.columns:
        df["mensaje_sistema"] = [
            str(m).strip() if pd.notna(m) else "" for m in df["mensaje_sistema"]
        ]
    return df


def get_connection(db_path: str = None):
    """
    Devuelve la conexión a usar. Si hay credenciales de Supabase configuradas
    (SUPABASE_DB_URL), siempre se conecta ahí — db_path queda sin efecto a
    propósito, porque con Supabase hay una sola base compartida entre la web
    y la app de escritorio, no un archivo por app. Sin esa variable, se
    comporta como siempre (SQLite local en db_path).

    Se vuelve a resolver la URL en cada llamada (no se usa solo la constante
    de módulo) porque en Streamlit Cloud el secret puede no estar disponible
    todavía en el instante en que se importa este módulo; al resolverlo acá,
    en tiempo de ejecución, ya está poblado. Ver _leer_supabase_url().
    """
    url = _leer_supabase_url()
    if url:
        return supabase_pg.SupabaseConnection(url)
    db_path = db_path or DB_PATH
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ------------------------------------------------------- caché de lecturas
#
# Streamlit re-ejecuta TODO el script en cada interacción (cada tecla del
# buscador, cada clic, cada cambio de pestaña) y, con st.tabs, además corren
# los 11 paneles del encargado en cada corrida. Sin caché eso repetía las
# mismas consultas pesadas contra la base remota decenas de veces: cargar el
# índice completo de alias en cada tecla, releer todo el inventario al pintar
# cada panel, etc. Contra Supabase (~85 ms por consulta) eso era el ~1.5 s de
# lag por acción.
#
# Se cachean SOLO lecturas de datos de REFERENCIA que cambian poco (catálogo
# de productos e índice de alias del buscador). Los datos transaccionales
# —solicitudes activas, correos autorizados, historial— NO se cachean: tienen
# que verse frescos al instante (una solicitud nueva debe aparecer apenas se
# crea) y su consulta es liviana igual.
#
# El caché es un TTL corto por proceso: dentro de una ráfaga de clics se reusa
# lo ya cargado, y se refresca solo cada tanto. Además se invalida explícita-
# mente cuando el encargado toca el catálogo, los alias o el stock (ver
# invalidar_cache_lecturas), para que esos cambios se vean sin esperar el TTL.
_CACHE_TTL_SEG = 20
_cache_lecturas = {}


def _leer_cacheado(clave, productor, ttl=_CACHE_TTL_SEG):
    ahora = time.monotonic()
    entrada = _cache_lecturas.get(clave)
    if entrada is not None and ahora - entrada[0] < ttl:
        return entrada[1]
    valor = productor()
    _cache_lecturas[clave] = (ahora, valor)
    return valor


def invalidar_cache_lecturas():
    """
    Vacía el caché de datos de referencia. Se llama tras cualquier escritura
    que cambie el catálogo, los alias o el saldo de un producto (cargar
    catálogo, crear/editar/borrar alias, cerrar una solicitud, importar
    saldos), para que el cambio se vea de inmediato en el buscador y en el
    inventario sin esperar a que expire el TTL.
    """
    _cache_lecturas.clear()


def generar_folio(prefijo="SOL") -> str:
    """
    OBSOLETO (día 4.5). El folio ahora es "Solicitud-<correlativo>" (ver
    crear_solicitud), que es el mismo número con que la municipalidad archiva
    el papel. Se conserva esta función solo para no romper bases antiguas que
    ya tengan folios con el formato viejo SOL-fecha-hash.
    """
    return f"{prefijo}-{ahora_chile().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"


def formatear_cantidad(valor) -> int:
    """Los productos de bodega se cuentan en unidades enteras (cajas, resmas,
    unidades, etc.) — no hay medio clip ni media resma. Se redondea y se
    muestra siempre como entero.

    Trata None y NaN por igual (devuelve 0): una cantidad_entregada aún NULL
    la lee pandas como NaN, e `int(round(nan))` reventaba con ValueError. Con
    pd.isna se cubren None, NaN y pd.NA sin que crashee."""
    if valor is None or pd.isna(valor):
        return 0
    return int(round(float(valor)))


# --------------------------------------------------------------- esquema BD

# Separado de init_db() para poder crear el mismo esquema en un archivo
# SQLite local directo (ver respaldar_base_local), sin pasar por
# get_connection() — que con Supabase configurado ignora a propósito cualquier
# db_path y siempre apunta a la base remota compartida.
ESQUEMA_SQL = \
        """
        CREATE TABLE IF NOT EXISTS productos (
            codigo TEXT PRIMARY KEY,
            nombre_estandar TEXT NOT NULL,
            unidad_medida TEXT,
            categoria TEXT,
            saldo REAL DEFAULT 0,
            saldo_importado REAL,
            fecha_corte TEXT,
            stock_critico REAL DEFAULT 0,
            valor_saldo REAL DEFAULT 0,
            valor_unitario REAL DEFAULT 0,
            ubicacion TEXT,
            fecha_venc TEXT,
            lote TEXT,
            activo INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS alias_productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            texto_alias TEXT NOT NULL,
            texto_alias_normalizado TEXT NOT NULL,
            codigo_producto TEXT NOT NULL,
            FOREIGN KEY (codigo_producto) REFERENCES productos(codigo)
        );

        CREATE TABLE IF NOT EXISTS alias_pendientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            texto_ingresado TEXT NOT NULL,
            fecha TEXT NOT NULL,
            revisado INTEGER DEFAULT 0
        );

        -- Personas registradas con correo institucional. El registro se hace
        -- una sola vez; desde ahí en adelante sus solicitudes quedan
        -- asociadas automáticamente a su nombre/área/supervisor, sin que
        -- tengan que volver a escribirlos (y sin que puedan inventarlos).
        -- Correos institucionales autorizados a hacer solicitudes.
        -- El dominio correcto ya no basta: alguien puede inventar un correo
        -- del dominio municipal que no existe. Solo los correos que el
        -- encargado cargue/autorice aquí pueden registrarse.
        CREATE TABLE IF NOT EXISTS correos_autorizados (
            correo TEXT PRIMARY KEY,
            nombre_referencia TEXT,
            area_departamento TEXT,
            estado TEXT DEFAULT 'autorizado',  -- autorizado | bloqueado
            fecha_alta TEXT NOT NULL,
            dado_de_alta_por TEXT,
            rol_al_registrar TEXT  -- rol que tomará la cuenta al registrarse (admin/encargado); NULL = solicitante
        );

        CREATE TABLE IF NOT EXISTS personas_registradas (
            correo TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            area_departamento TEXT NOT NULL,
            nombre_supervisor TEXT,
            correo_supervisor TEXT,
            password_hash TEXT,
            password_salt TEXT,
            rol TEXT DEFAULT 'solicitante',
            fecha_registro TEXT NOT NULL
        );
        -- Alias que un solicitante "sugiere" al elegir un match <100%.
        -- NO se activa solo: el encargado debe aprobarlo para que quede
        -- como alias_productos real y sirva en futuras búsquedas.
        CREATE TABLE IF NOT EXISTS alias_sugeridos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            texto_alias TEXT NOT NULL,
            texto_alias_normalizado TEXT NOT NULL,
            codigo_producto TEXT NOT NULL,
            score REAL,
            fecha TEXT NOT NULL,
            estado TEXT DEFAULT 'pendiente',  -- pendiente | aprobado | rechazado
            FOREIGN KEY (codigo_producto) REFERENCES productos(codigo)
        );

        -- Cabecera de solicitud (folio = liga el papel firmado con el registro digital)
        CREATE TABLE IF NOT EXISTS solicitudes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            folio TEXT UNIQUE NOT NULL,
            fecha_solicitud TEXT NOT NULL,
            solicitante TEXT,
            supervisor TEXT,
            area_departamento TEXT,
            estado TEXT DEFAULT 'pendiente_firma',
            -- pendiente_firma -> preliminar_aceptada -> editada -> cerrada
            --                                        -> anulada (en cualquier punto antes de cerrada)
            sincronizado_smc INTEGER DEFAULT 0,
            motivo_anulacion TEXT,
            correo_solicitante TEXT,
            correo_supervisor TEXT,
            correlativo INTEGER,
            oficina TEXT,
            usuario_operacion TEXT,
            info_adicional TEXT,
            depto_origen TEXT
        );

        -- Parámetros que el encargado ajusta desde la interfaz y deben
        -- sobrevivir a reinicios (ej. su nombre y apellido, que va impreso
        -- en el comprobante como responsable del movimiento).
        CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT
        );

        -- Detalle: uno o más productos por solicitud
        CREATE TABLE IF NOT EXISTS solicitud_detalle (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitud_id INTEGER NOT NULL,
            codigo_producto TEXT NOT NULL,
            cantidad_solicitada REAL NOT NULL,
            cantidad_entregada REAL,
            mensaje_sistema TEXT,
            valor_movimiento REAL,  -- valor en pesos de lo entregado (valor_unitario * cantidad), fijado al cerrar
            FOREIGN KEY (solicitud_id) REFERENCES solicitudes(id),
            FOREIGN KEY (codigo_producto) REFERENCES productos(codigo)
        );

        CREATE TABLE IF NOT EXISTS alertas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_producto TEXT,
            tipo TEXT,  -- 'agotado' | 'critico'
            mensaje TEXT,
            fecha TEXT
        );

        -- Registro de corridas del job de sincronización hacia SMC.
        -- Hoy SMC no tiene una vía de integración confirmada, así que este
        -- job deja un archivo de intercambio (ver sincronizar_smc.py) en vez
        -- de escribir directo a otra base de datos. Si en el futuro se logra
        -- acceso real a SMC (archivo de importación, ODBC, etc.), este mismo
        -- registro sirve para saber qué falta enviar y qué ya se envió.
        CREATE TABLE IF NOT EXISTS log_sincronizacion_smc (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            folios_incluidos INTEGER,
            archivo_generado TEXT,
            estado TEXT  -- 'exportado_local' | 'enviado_smc' (hipotético, no usado hoy)
        );
        """


def init_db(db_path: str = None) -> None:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.executescript(ESQUEMA_SQL)
    conn.commit()

    # Migración defensiva: si esta función corre sobre un bodega.db creado
    # con una versión anterior (antes de 'día 2'), las tablas ya existen y
    # CREATE TABLE IF NOT EXISTS no les agrega las columnas nuevas. Se
    # intentan agregar aquí; si ya existen, sqlite tira OperationalError y
    # simplemente se ignora.
    migraciones = [
        "ALTER TABLE solicitudes ADD COLUMN sincronizado_smc INTEGER DEFAULT 0",
        "ALTER TABLE solicitudes ADD COLUMN motivo_anulacion TEXT",
        "ALTER TABLE productos ADD COLUMN saldo_importado REAL",
        "ALTER TABLE productos ADD COLUMN valor_saldo REAL DEFAULT 0",
        "ALTER TABLE productos ADD COLUMN fecha_corte TEXT",
        "ALTER TABLE solicitudes ADD COLUMN correo_solicitante TEXT",
        "ALTER TABLE solicitudes ADD COLUMN correo_supervisor TEXT",
        "ALTER TABLE personas_registradas ADD COLUMN nombre_supervisor TEXT",
        "ALTER TABLE solicitudes ADD COLUMN correlativo INTEGER",
        "ALTER TABLE solicitudes ADD COLUMN oficina TEXT",
        "ALTER TABLE solicitudes ADD COLUMN usuario_operacion TEXT",
        "ALTER TABLE solicitudes ADD COLUMN info_adicional TEXT",
        "ALTER TABLE solicitudes ADD COLUMN depto_origen TEXT",
        "ALTER TABLE personas_registradas ADD COLUMN password_hash TEXT",
        "ALTER TABLE personas_registradas ADD COLUMN password_salt TEXT",
        "ALTER TABLE personas_registradas ADD COLUMN rol TEXT DEFAULT 'solicitante'",
        "ALTER TABLE solicitudes ADD COLUMN memo TEXT",
        "ALTER TABLE solicitudes ADD COLUMN tipo_movimiento TEXT",
        "ALTER TABLE solicitudes ADD COLUMN destino TEXT",
        "ALTER TABLE productos ADD COLUMN valor_unitario REAL DEFAULT 0",
        "ALTER TABLE solicitud_detalle ADD COLUMN valor_movimiento REAL",
        "ALTER TABLE correos_autorizados ADD COLUMN rol_al_registrar TEXT",
    ]
    for sql in migraciones:
        try:
            conn.execute(sql)
            conn.commit()
        except (sqlite3.OperationalError, supabase_pg.SupabaseError):
            pass  # la columna ya existía

    conn.close()


def catalogo_cargado(db_path: str = None) -> bool:
    """
    True si la tabla productos ya tiene datos. Reemplaza a comprobar si el
    archivo de base existe: con Supabase, DB_PATH es un connection string
    (siempre "existe" para os.path.exists, que ni siquiera aplica ahí), así
    que la única forma confiable de saber si hace falta cargar_catalogo() es
    preguntarle a la base directamente.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    try:
        n = conn.execute("SELECT COUNT(*) FROM productos").fetchone()[0]
    except (sqlite3.OperationalError, supabase_pg.SupabaseError):
        n = 0
    conn.close()
    return n > 0


TABLAS_RESPALDABLES = (
    "productos", "alias_productos", "alias_pendientes", "correos_autorizados",
    "personas_registradas", "alias_sugeridos", "solicitudes", "configuracion",
    "solicitud_detalle", "alertas", "log_sincronizacion_smc",
)


def respaldar_base_local(ruta_destino, db_path: str = None) -> None:
    """
    Guarda una FOTO de la base actual (Supabase) en un archivo SQLite en el
    propio equipo — no es una réplica en vivo ni se usa para leer: la app
    siempre lee y escribe contra Supabase. Sirve solo como resguardo "por si
    acaso" ante un corte de internet o un problema con el servicio. Cada
    llamada sobrescribe el archivo anterior con el estado más reciente.
    """
    ruta_destino = str(ruta_destino)
    if os.path.exists(ruta_destino):
        os.remove(ruta_destino)

    # sqlite3 directo, no get_connection(): con Supabase configurado,
    # get_connection() ignora a propósito cualquier ruta local (ver su
    # docstring) — acá se necesita sí o sí un archivo en disco.
    destino = sqlite3.connect(ruta_destino)
    destino.executescript(ESQUEMA_SQL)

    origen = get_connection(db_path)
    for tabla in TABLAS_RESPALDABLES:
        filas = origen.execute(f"SELECT * FROM {tabla}").fetchall()
        if not filas:
            continue
        n_columnas = len(filas[0])
        placeholders = ", ".join(["?"] * n_columnas)
        destino.executemany(
            f"INSERT INTO {tabla} VALUES ({placeholders})", filas,
        )
    destino.commit()
    destino.close()
    origen.close()


# --------------------------------------------------------- carga de catálogo

def cargar_catalogo(productos, db_path: str = None):
    """
    productos: lista de tuplas (codigo, nombre, unidad, saldo, stock_critico, valor_saldo).
    valor_saldo es el valor contable total de esa línea tal como aparece impreso
    en la columna 'SALDO VAL.' del listado real. De ese total y la cantidad del
    corte se deriva UNA vez el valor unitario (ver valor_unitario_desde): es el
    "momento del escaneo", el único punto donde se fija ese precio por unidad,
    que después se usa fijo para mover el inventario en pesos con cada entrega.
    """
    db_path = db_path or DB_PATH
    from catalogo_real import categoria_por_codigo

    conn = get_connection(db_path)
    for producto in productos:
        codigo, nombre, unidad, saldo, stock_critico, valor_saldo = producto
        unitario = valor_unitario_desde(valor_saldo, saldo)
        conn.execute(
            """
            INSERT INTO productos
                (codigo, nombre_estandar, unidad_medida, categoria, saldo, stock_critico,
                 valor_saldo, valor_unitario, activo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            ON CONFLICT (codigo) DO UPDATE SET
                nombre_estandar = EXCLUDED.nombre_estandar,
                unidad_medida = EXCLUDED.unidad_medida,
                categoria = EXCLUDED.categoria,
                saldo = EXCLUDED.saldo,
                stock_critico = EXCLUDED.stock_critico,
                valor_saldo = EXCLUDED.valor_saldo,
                valor_unitario = EXCLUDED.valor_unitario,
                activo = EXCLUDED.activo
            """,
            (codigo, nombre, unidad, categoria_por_codigo(codigo), saldo, stock_critico,
             valor_saldo, unitario),
        )
        # el nombre estándar también queda registrado como su propio alias de búsqueda
        texto_norm = normalizar(nombre)
        conn.execute(
            "INSERT INTO alias_productos (texto_alias, texto_alias_normalizado, codigo_producto) "
            "VALUES (?, ?, ?)",
            (nombre, texto_norm, codigo),
        )
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()


# ------------------------------------------------------------ tabla de alias

def registrar_alias_nuevo(texto_alias, codigo_producto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    texto_norm = normalizar(texto_alias)
    existe = conn.execute(
        "SELECT 1 FROM alias_productos WHERE texto_alias_normalizado=? AND codigo_producto=?",
        (texto_norm, codigo_producto),
    ).fetchone()
    if not existe:
        conn.execute(
            "INSERT INTO alias_productos (texto_alias, texto_alias_normalizado, codigo_producto) "
            "VALUES (?, ?, ?)",
            (texto_alias, texto_norm, codigo_producto),
        )
        conn.commit()
        invalidar_cache_lecturas()
    conn.close()


def obtener_producto(codigo_producto, db_path: str = None):
    """
    Devuelve los datos del producto para corroborar que el código existe
    ANTES de asociarle un alias. Devuelve None si el código no está en el
    catálogo.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT codigo, nombre_estandar, unidad_medida, categoria, saldo "
        "FROM productos WHERE codigo = ? AND activo = 1",
        (str(codigo_producto).strip(),),
    ).fetchone()
    conn.close()
    if fila is None:
        return None
    return {
        "codigo": fila[0], "nombre_estandar": fila[1], "unidad_medida": fila[2],
        "categoria": fila[3], "saldo": formatear_cantidad(fila[4]),
    }


def listar_alias_de_producto(codigo_producto, db_path: str = None):
    """Alias ya registrados para un producto (para no duplicar al crear uno nuevo)."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    filas = conn.execute(
        "SELECT texto_alias FROM alias_productos WHERE codigo_producto = ? ORDER BY id",
        (str(codigo_producto).strip(),),
    ).fetchall()
    conn.close()
    return [f[0] for f in filas]


def alias_de_producto(codigo_producto, db_path: str = None) -> pd.DataFrame:
    """Alias de un producto con su id, para poder editarlos o borrarlos."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        "SELECT id, texto_alias, texto_alias_normalizado FROM alias_productos "
        "WHERE codigo_producto = ? ORDER BY id",
        conn, params=(str(codigo_producto).strip(),),
    )
    conn.close()
    return df


def listar_alias_para_gestion(filtro="", solo_manuales=True, limite=300,
                              db_path: str = None) -> pd.DataFrame:
    """
    Alias registrados, con su producto, para revisarlos y corregirlos SIN
    tener que saber el código de cada producto.

    Por defecto (solo_manuales=True) excluye el alias que es igual al nombre
    estándar del producto: ese se crea solo al cargar el catálogo (cada
    producto queda buscable por su propio nombre) y no tiene sentido tocarlo
    acá. Quedan así a la vista los alias que agregó el encargado a mano, que
    son los que pueden estar mal escritos o apuntando al producto equivocado.

    filtro: si se entrega, filtra por texto del alias O nombre del producto
    (sin distinguir mayúsculas). Vacío = todos (hasta 'limite').
    """
    db_path = db_path or DB_PATH
    condiciones = ["p.activo = 1"]
    params = []
    if solo_manuales:
        # el alias auto-generado se insertó con texto_alias EXACTAMENTE igual
        # al nombre estándar (ver cargar_catalogo)
        condiciones.append("a.texto_alias <> p.nombre_estandar")
    filtro = (filtro or "").strip().lower()
    if filtro:
        condiciones.append("(LOWER(a.texto_alias) LIKE ? OR LOWER(p.nombre_estandar) LIKE ?)")
        params.extend([f"%{filtro}%", f"%{filtro}%"])
    where = " AND ".join(condiciones)
    params.append(int(limite))

    conn = get_connection(db_path)
    df = pd.read_sql(
        f"""
        SELECT a.id, a.texto_alias, a.codigo_producto, p.nombre_estandar
        FROM alias_productos a
        JOIN productos p ON p.codigo = a.codigo_producto
        WHERE {where}
        ORDER BY p.nombre_estandar, a.id
        LIMIT ?
        """,
        conn, params=tuple(params),
    )
    conn.close()
    return df


def eliminar_alias(id_alias, db_path: str = None):
    """
    Borra un alias. No deja al producto sin ninguna forma de ser encontrado:
    si es el último que le queda, se rechaza el borrado, porque un producto
    sin alias desaparece del buscador y nadie podría volver a pedirlo.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT codigo_producto, texto_alias FROM alias_productos WHERE id = ?", (id_alias,)
    ).fetchone()
    if fila is None:
        conn.close()
        return False, "Ese alias ya no existe."
    codigo, texto = fila
    total = conn.execute(
        "SELECT COUNT(*) FROM alias_productos WHERE codigo_producto = ?", (codigo,)
    ).fetchone()[0]
    if total <= 1:
        conn.close()
        return False, ("No se puede borrar: es el único alias del producto y quedaría "
                       "imposible de encontrar en el buscador. Cree otro antes de borrar este.")
    conn.execute("DELETE FROM alias_productos WHERE id = ?", (id_alias,))
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()
    return True, f'Alias "{texto}" eliminado.'


def editar_alias(id_alias, nuevo_texto, db_path: str = None):
    """Corrige el texto de un alias existente (ej. un error de tipeo)."""
    db_path = db_path or DB_PATH
    nuevo_texto = (nuevo_texto or "").strip()
    if not nuevo_texto:
        return False, "El alias no puede quedar vacío."

    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT codigo_producto FROM alias_productos WHERE id = ?", (id_alias,)
    ).fetchone()
    if fila is None:
        conn.close()
        return False, "Ese alias ya no existe."
    codigo = fila[0]
    norm = normalizar(nuevo_texto)

    conflicto = conn.execute(
        """
        SELECT a.codigo_producto, p.nombre_estandar
        FROM alias_productos a JOIN productos p ON p.codigo = a.codigo_producto
        WHERE a.texto_alias_normalizado = ? AND a.id != ?
        """,
        (norm, id_alias),
    ).fetchone()
    if conflicto:
        conn.close()
        if conflicto[0] == codigo:
            return False, f'"{nuevo_texto}" ya está registrado para este mismo producto.'
        return False, (f'"{nuevo_texto}" ya apunta a otro producto: {conflicto[1]} '
                       f"(código {conflicto[0]}).")

    conn.execute(
        "UPDATE alias_productos SET texto_alias=?, texto_alias_normalizado=? WHERE id=?",
        (nuevo_texto, norm, id_alias),
    )
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()
    return True, f'Alias actualizado a "{nuevo_texto}".'


def crear_alias_manual(codigo_producto, texto_alias, db_path: str = None):
    """
    Crea un alias de búsqueda para un producto EXISTENTE. Es la operación que
    hace el encargado a mano: no cambia el nombre del producto, solo agrega
    una forma más en que la gente lo puede escribir en el buscador
    (ej. 'confort' -> ROLLO PAPEL HIGIÉNICO; 'poet' -> LIMPIADOR MULTIUSO).

    Devuelve (ok: bool, mensaje: str).
    """
    db_path = db_path or DB_PATH
    codigo = str(codigo_producto).strip()
    alias = (texto_alias or "").strip()

    if not alias:
        return False, "El alias no puede estar vacío."

    producto = obtener_producto(codigo, db_path)
    if producto is None:
        return False, f'El código "{codigo}" no existe en el catálogo. Verifique el código antes de crear el alias.'

    alias_norm = normalizar(alias)

    # ¿este alias ya apunta a otro producto? Es importante avisarlo: un mismo
    # texto apuntando a dos productos distintos vuelve ambiguo el buscador.
    conn = get_connection(db_path)
    conflicto = conn.execute(
        """
        SELECT a.codigo_producto, p.nombre_estandar
        FROM alias_productos a JOIN productos p ON p.codigo = a.codigo_producto
        WHERE a.texto_alias_normalizado = ? AND a.codigo_producto != ?
        """,
        (alias_norm, codigo),
    ).fetchone()
    ya_existe = conn.execute(
        "SELECT 1 FROM alias_productos WHERE texto_alias_normalizado = ? AND codigo_producto = ?",
        (alias_norm, codigo),
    ).fetchone()
    conn.close()

    if ya_existe:
        return False, f'"{alias}" ya estaba registrado como alias de {producto["nombre_estandar"]}.'
    if conflicto:
        return False, (f'"{alias}" ya apunta a otro producto: {conflicto[1]} (código {conflicto[0]}). '
                       f"Use un alias distinto o revise cuál de los dos corresponde.")

    registrar_alias_nuevo(alias, codigo, db_path)
    return True, f'Alias "{alias}" creado para {producto["nombre_estandar"]} (código {codigo}).'


def importar_alias_desde_excel(ruta_archivo, db_path: str = None) -> pd.DataFrame:
    """
    Carga masiva de alias desde un Excel/CSV que el encargado llena a mano.

    El archivo debe tener dos columnas, llamadas 'codigo' y 'alias'
    (no importan mayúsculas ni el orden de las columnas). Cada fila es
    "esta palabra que la gente escribe -> este código de producto".

    Procesa fila por fila y devuelve un DataFrame con el resultado de cada
    una (creado / rechazado y por qué), para que el encargado vea exactamente
    qué entró y qué no en vez de un "listo" a ciegas.
    """
    db_path = db_path or DB_PATH
    def _es_csv(r):
        return str(getattr(r, "name", r)).lower().endswith((".csv", ".txt"))

    def _leer(header):
        if _es_csv(ruta_archivo):
            return pd.read_csv(ruta_archivo, dtype=str, header=header)
        return pd.read_excel(ruta_archivo, dtype=str, header=header)

    ruta = ruta_archivo

    # La fila de títulos no siempre es la primera: la plantilla trae un
    # encabezado con instrucciones arriba, y el encargado podría agregar
    # sus propias notas. Se busca la fila que contenga 'codigo' y 'alias'
    # en las primeras 20 filas, en vez de asumir que es la fila 1.
    crudo = _leer(None)
    fila_encabezado = None
    for i in range(min(20, len(crudo))):
        valores = [str(v).strip().lower() for v in crudo.iloc[i].tolist()]
        if "codigo" in valores and "alias" in valores:
            fila_encabezado = i
            break

    if fila_encabezado is None:
        raise ValueError(
            "No se encontró una fila de títulos con las columnas 'codigo' y 'alias' "
            "en las primeras filas del archivo. Revise que existan esas dos columnas."
        )

    if hasattr(ruta_archivo, "seek"):
        ruta_archivo.seek(0)  # archivo subido por Streamlit: rebobinar antes de releer
    df = _leer(fila_encabezado)
    df.columns = [str(c).strip().lower() for c in df.columns]

    resultados = []
    for _, fila in df.iterrows():
        codigo = (fila.get("codigo") or "").strip() if isinstance(fila.get("codigo"), str) else ""
        alias = (fila.get("alias") or "").strip() if isinstance(fila.get("alias"), str) else ""
        if not codigo and not alias:
            continue  # fila vacía de relleno, se ignora en silencio
        ok, mensaje = crear_alias_manual(codigo, alias, db_path)
        resultados.append({
            "codigo": codigo, "alias": alias,
            "resultado": "creado" if ok else "rechazado",
            "detalle": mensaje,
        })

    return pd.DataFrame(resultados)


def sugerir_alias(texto_alias, codigo_producto, score, db_path: str = None):
    """
    Usado por la interfaz del SOLICITANTE cuando elige un match <100%.
    A diferencia de registrar_alias_nuevo(), esto NO activa el alias de
    inmediato: queda 'pendiente' hasta que el encargado lo apruebe desde
    su propia interfaz. Un solicitante nunca puede crear un alias real por
    su cuenta.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    texto_norm = normalizar(texto_alias)
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")
    ya_sugerido = conn.execute(
        "SELECT 1 FROM alias_sugeridos WHERE texto_alias_normalizado=? AND codigo_producto=? AND estado='pendiente'",
        (texto_norm, codigo_producto),
    ).fetchone()
    ya_alias = conn.execute(
        "SELECT 1 FROM alias_productos WHERE texto_alias_normalizado=? AND codigo_producto=?",
        (texto_norm, codigo_producto),
    ).fetchone()
    if not ya_sugerido and not ya_alias:
        conn.execute(
            "INSERT INTO alias_sugeridos (texto_alias, texto_alias_normalizado, codigo_producto, score, fecha) "
            "VALUES (?, ?, ?, ?, ?)",
            (texto_alias, texto_norm, codigo_producto, score, fecha),
        )
        conn.commit()
    conn.close()


def listar_alias_sugeridos(estado="pendiente", db_path: str = None) -> pd.DataFrame:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.id, s.texto_alias, s.codigo_producto, p.nombre_estandar, s.score, s.fecha, s.estado
        FROM alias_sugeridos s
        JOIN productos p ON p.codigo = s.codigo_producto
        WHERE s.estado = ?
        ORDER BY s.id DESC
        """,
        conn,
        params=(estado,),
    )
    conn.close()
    return df


def aprobar_alias_sugerido(id_sugerencia, db_path: str = None):
    """Solo el encargado ejecuta esto. Activa el alias definitivamente."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT texto_alias, codigo_producto FROM alias_sugeridos WHERE id=?", (id_sugerencia,)
    ).fetchone()
    conn.close()
    if fila:
        texto_alias, codigo_producto = fila
        registrar_alias_nuevo(texto_alias, codigo_producto, db_path)
        conn = get_connection(db_path)
        conn.execute("UPDATE alias_sugeridos SET estado='aprobado' WHERE id=?", (id_sugerencia,))
        conn.commit()
        conn.close()


def rechazar_alias_sugerido(id_sugerencia, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE alias_sugeridos SET estado='rechazado' WHERE id=?", (id_sugerencia,))
    conn.commit()
    conn.close()


def registrar_alias_pendiente(texto_ingresado, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")
    conn.execute(
        "INSERT INTO alias_pendientes (texto_ingresado, fecha) VALUES (?, ?)",
        (texto_ingresado, fecha),
    )
    conn.commit()
    conn.close()


def _puntaje(consulta_norm: str, candidato_norm: str, *, score_cutoff=None, **_kwargs) -> float:
    """
    Puntaje de coincidencia mostrado al usuario.

    Se usa una mezcla de dos scorers porque cada uno solo falla:
      - token_set_ratio devuelve 100 cuando lo buscado es subconjunto del
        nombre, así que "lapiz" daba 100% a LAPIZ PASTA AZUL, LAPIZ PASTA
        ROJO, LAPIZ PASTA NEGRO... y "100% = coincidencia exacta" perdía
        todo significado en pantalla.
      - token_sort_ratio castiga las palabras de más, pero por sí solo
        hunde coincidencias correctas cuando el nombre del catálogo es
        largo (que es la norma acá: "BOLSAS DE BASURA 70 x 90 CM X 10
        UNIDADES").
    La mezcla mantiene el buen ordenamiento del primero y recupera la
    gradación del segundo, de modo que el 100% queda reservado para lo que
    de verdad calza entero.
    """
    base = (0.55 * fuzz.token_set_ratio(consulta_norm, candidato_norm) +
            0.45 * fuzz.token_sort_ratio(consulta_norm, candidato_norm))

    # Refuerzo por coincidencia de palabra completa.
    #
    # Sin esto, buscar "tinta" mostraba solo las TINTA TIMBRE: como
    # token_sort_ratio castiga las palabras de más, ganaban siempre los
    # nombres más cortos y las ~60 tintas restantes quedaban fuera del corte.
    # Si lo escrito aparece como palabra al inicio del nombre (o dentro de
    # él), es casi seguro que el producto es de la familia buscada, así que
    # se le sube el puntaje. Se topa en 99 para que el 100 siga estando
    # reservado a la coincidencia exacta.
    palabras = candidato_norm.split()
    consulta_palabras = consulta_norm.split()
    if not consulta_palabras:
        return base
    if candidato_norm.startswith(consulta_norm):
        return min(99, max(base, 88))
    if all(p in palabras for p in consulta_palabras):
        return min(99, max(base, 78))
    return base


def buscar_producto(texto_busqueda: str, db_path: str = None, limite: int = 15, umbral: int = 45):
    """
    Devuelve candidatos (codigo, nombre_estandar, score) para un texto libre,
    buscando primero coincidencia exacta en los alias y luego aproximada.

    NOTA TÉCNICA (corrección día 2):
    La versión anterior usaba fuzz.WRatio, que en rapidfuzz pondera fuerte el
    "partial_ratio" (coincidencia de subcadenas). Con nombres de producto largos
    eso hace que términos cortos como "grapas" o "papel confort" reciban el
    mismo puntaje (85%-90%) contra productos sin ninguna relación real
    (ej. "grapas" -> "PASTILLA PARA ESTANQUE"), y que el producto correcto
    (ej. "BOLSAS DE BASURA" al buscar "bolsa basura") quede empatado o por
    debajo de coincidencias irrelevantes. Se cambió a fuzz.token_set_ratio,
    que compara conjuntos de palabras (ignora orden y palabras de más) y
    diferencia mucho mejor la relevancia real. Además se agregó un desempate
    por cercanía de longitud y se bajó el umbral pero de forma más estricta
    en la práctica, porque token_set_ratio ya no infla puntajes artificialmente.
    """
    db_path = db_path or DB_PATH

    # El índice de alias (una fila por forma de nombrar cada producto) es lo
    # que se recarga en cada tecla del buscador. Se cachea porque cambia solo
    # cuando el encargado agrega/edita alias — momento en que se invalida.
    def _cargar_indice_alias():
        conn = get_connection(db_path)
        df = pd.read_sql(
            """
            SELECT a.texto_alias_normalizado, a.codigo_producto, p.nombre_estandar
            FROM alias_productos a
            JOIN productos p ON a.codigo_producto = p.codigo
            WHERE p.activo = 1
            """,
            conn,
        )
        conn.close()
        return df

    df_alias = _leer_cacheado(("indice_alias", db_path), _cargar_indice_alias)
    if df_alias.empty:
        return []

    texto_norm = normalizar(texto_busqueda)

    # Coincidencia exacta: va primero con 100%, pero NO corta la búsqueda.
    #
    # Antes se devolvía solo ese producto y se descartaba todo lo demás. Eso
    # provocaba un efecto raro: si el encargado buscaba "tinta canon" y elegía
    # la cyan, se creaba el alias "tinta canon" apuntando a ella; en la
    # búsqueda siguiente ese alias calzaba exacto y las otras tintas canon
    # desaparecían de la lista. El 100% (la estrella) sirve para confirmar que
    # hay una coincidencia segura, no para anular el resto de las opciones.
    exactos = df_alias[df_alias["texto_alias_normalizado"] == texto_norm]
    candidatos, vistos = [], set()
    for _, fila in exactos.iterrows():
        codigo = fila["codigo_producto"]
        if codigo not in vistos:
            candidatos.append((codigo, fila["nombre_estandar"], 100))
            vistos.add(codigo)

    opciones = df_alias["texto_alias_normalizado"].tolist()
    resultados = process.extract(texto_norm, opciones, scorer=_puntaje, limit=limite * 6)

    # Desempate: mayor score primero; en empate, el nombre de largo más
    # parecido al texto buscado (evita que gane un producto larguísimo
    # que "contiene" las palabras pero es otra cosa).
    resultados.sort(key=lambda r: (-r[1], abs(len(r[0]) - len(texto_norm))))

    aproximados = []
    for _texto, score, idx in resultados:
        fila = df_alias.iloc[idx]
        codigo = fila["codigo_producto"]
        if codigo not in vistos and score >= umbral:
            aproximados.append((codigo, fila["nombre_estandar"], round(score, 1)))
            vistos.add(codigo)

    candidatos.extend(_diversificar(aproximados, limite - len(candidatos)))
    return candidatos[:limite]


def _familia(nombre: str) -> str:
    """
    Segunda palabra del nombre, que en este catálogo suele ser la marca o el
    tipo: "TINTA CANON ...", "TINTA EPSON ...", "TINTA TIMBRE ...".
    Sirve para no llenar la lista con una sola familia.
    """
    palabras = normalizar(nombre).split()
    return palabras[1] if len(palabras) > 1 else (palabras[0] if palabras else "")


def _diversificar(candidatos, cupo):
    """
    Reparte los cupos entre familias, pero SOLO entre productos que empatan en
    puntaje. Nunca deja que uno peor se cuele delante de uno mejor.

    Sin esto, buscar "tinta" devolvía 15 resultados con el mismo puntaje pero
    todos de dos o tres familias (TIMBRE, HP, EPSON), y las tintas CANON no
    aparecían nunca aunque coincidieran igual de bien. Y si se diversifica sin
    respetar el puntaje pasa lo contrario: al buscar "bolsa basura" se colaban
    productos irrelevantes y desaparecían dos de las tres bolsas reales.
    """
    if cupo <= 0:
        return []

    # se agrupa por puntaje (redondeado) manteniendo el orden de mejor a peor
    por_puntaje = {}
    for candidato in candidatos:
        por_puntaje.setdefault(round(candidato[2]), []).append(candidato)

    seleccion = []
    for puntaje in sorted(por_puntaje, reverse=True):
        grupo = por_puntaje[puntaje]
        if len(seleccion) >= cupo:
            break

        # dentro del empate, una de cada familia por vuelta
        familias = {}
        for candidato in grupo:
            familias.setdefault(_familia(candidato[1]), []).append(candidato)

        vuelta = 0
        while len(seleccion) < cupo:
            agregado = False
            for familia in familias:
                if vuelta < len(familias[familia]):
                    seleccion.append(familias[familia][vuelta])
                    agregado = True
                    if len(seleccion) >= cupo:
                        break
            if not agregado:
                break
            vuelta += 1

    return seleccion


# --------------------------------------------- validación (etapa de especificación)

def validar_disponibilidad(codigo_producto, cantidad_solicitada, db_path: str = None):
    """
    Reglas del diagnóstico inicial:
      insumo == 0          -> "no cuenta con insumos registrados en sistema"
      insumo < solicitado  -> "no se encuentra la cantidad especificada, hay x"
      insumo >= solicitado -> disponible, sin mensaje
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    row = conn.execute(
        "SELECT saldo, nombre_estandar, fecha_corte FROM productos WHERE codigo=?",
        (codigo_producto,),
    ).fetchone()
    conn.close()
    if row is None:
        return "error", "Producto no encontrado en catálogo."
    saldo, nombre, corte = row
    saldo_i = formatear_cantidad(saldo)
    # El saldo es una estimación desde el último corte de SMC, no un dato en
    # vivo: se dice de dónde viene para que nadie lo tome como definitivo.
    referencia = f" (saldo al corte del {str(corte)[:10]})" if corte else ""
    if saldo_i <= 0:
        return "agotado", (f'"{nombre}" no cuenta con insumos registrados en sistema'
                           f"{referencia}. Confirmar en bodega antes de asumir que está agotado.")
    if saldo_i < cantidad_solicitada:
        return "parcial", (f'En sistema no se encuentra la cantidad solicitada de "{nombre}"; '
                           f"hay {saldo_i} disponibles{referencia}.")
    return "ok", None


def evaluar_alerta_stock(codigo_producto, db_path: str = None):
    """
    Reglas de alerta para el operario, tras descontar stock:
      saldo == 0        -> agotado, recomendar verificar físicamente / recompra
      saldo < critico    -> alerta, recomendar recompra
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    row = conn.execute(
        "SELECT saldo, stock_critico, nombre_estandar FROM productos WHERE codigo=?",
        (codigo_producto,),
    ).fetchone()
    conn.close()
    saldo, stock_critico, nombre = row
    saldo_i = formatear_cantidad(saldo)
    critico_i = formatear_cantidad(stock_critico)
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")

    if saldo_i <= 0:
        msg = f'Se ha agotado el insumo "{nombre}". Se recomienda verificar físicamente y asignar recompra.'
        _guardar_alerta(codigo_producto, "agotado", msg, fecha, db_path)
        return "agotado", msg
    if critico_i and saldo_i < critico_i:
        msg = f'El insumo "{nombre}" quedó bajo su stock crítico ({critico_i}). Se recomienda recompra.'
        _guardar_alerta(codigo_producto, "critico", msg, fecha, db_path)
        return "critico", msg
    return "ok", None


def _guardar_alerta(codigo_producto, tipo, mensaje, fecha, db_path):
    conn = get_connection(db_path)
    conn.execute(
        "INSERT INTO alertas (codigo_producto, tipo, mensaje, fecha) VALUES (?, ?, ?, ?)",
        (codigo_producto, tipo, mensaje, fecha),
    )
    conn.commit()
    conn.close()


# -------------------------------------------- correos autorizados (whitelist)
#
# El dominio correcto no prueba que el correo exista: cualquiera puede
# escribir "juan.perez@<dominio-municipal>" sin ser esa persona. Como este
# sistema no puede enviar un correo de verificación (no hay servidor de
# correo conectado), el control real es una lista blanca que el encargado
# administra a mano: si el correo no está en la lista, no se puede registrar.

# Nombre y apellido de quien procesa las solicitudes. Va impreso en el
# comprobante como responsable del movimiento; se edita desde la interfaz
# del encargado (queda guardado en la tabla configuracion).
USUARIO_BODEGA_POR_DEFECTO = "Gonzalo Fierro Cea"

CORRELATIVO_INICIAL = 2900  # el talonario físico va en 2798 (mayo); se parte en 2900


# Partículas de enlace que dentro de un nombre/apellido van en minúscula (salvo
# cuando abren el nombre): "Juan de la Cruz", "María de los Ángeles".
_PARTICULAS_NOMBRE = {"de", "del", "la", "las", "los", "y", "e", "da", "do", "dos", "van", "von"}


def formatear_nombre_persona(nombre) -> str:
    """
    Deja un nombre de persona en un formato uniforme, para que no queden unos en
    MAYÚSCULAS, otros en minúscula y otros mezclados: inicial de cada palabra en
    mayúscula y el resto en minúscula, con los espacios colapsados. NO recorta
    palabras, así respeta apellidos compuestos ('SAN MIGUEL' -> 'San Miguel') y
    los segundos nombres/apellidos. Las partículas de enlace ('de', 'la',
    'del'...) quedan en minúscula salvo cuando abren el nombre.
    Ej.: 'cristian SAN miguel' -> 'Cristian San Miguel'.
    """
    palabras = str(nombre or "").split()
    resultado = []
    for i, palabra in enumerate(palabras):
        baja = palabra.lower()
        if i > 0 and baja in _PARTICULAS_NOMBRE:
            resultado.append(baja)
        else:
            # respeta guiones internos (Ana-María) capitalizando cada parte
            resultado.append("-".join(parte.capitalize() for parte in baja.split("-")))
    return " ".join(resultado)


def autorizar_correo(correo, nombre_referencia="", area_departamento="", dado_de_alta_por="encargado",
                     db_path: str = None, rol_al_registrar=None):
    """
    Autoriza un correo a registrarse. rol_al_registrar (opcional): si se pasa
    'admin' o 'encargado', la cuenta tomará ese rol al momento de registrarse
    (así se puede dar acceso de encargado/admin a alguien sin conocer su
    contraseña: se autoriza con el rol y la persona se registra de cero). Si
    es None, se registra como solicitante normal.
    """
    db_path = db_path or DB_PATH
    correo = (correo or "").strip().lower()
    if not correo:
        return False, "El correo no puede estar vacío."
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
        return False, f'"{correo}" no tiene formato de correo válido.'
    if rol_al_registrar not in (None, "admin", "encargado", "solicitante"):
        return False, "Rol a asignar inválido."
    if rol_al_registrar == "solicitante":
        rol_al_registrar = None  # es el rol por defecto: no hace falta guardarlo
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")
    conn = get_connection(db_path)
    conn.execute(
        "INSERT INTO correos_autorizados "
        "(correo, nombre_referencia, area_departamento, estado, fecha_alta, dado_de_alta_por, rol_al_registrar) "
        "VALUES (?, ?, ?, 'autorizado', ?, ?, ?) "
        "ON CONFLICT (correo) DO UPDATE SET "
        "nombre_referencia = EXCLUDED.nombre_referencia, "
        "area_departamento = EXCLUDED.area_departamento, "
        "estado = EXCLUDED.estado, "
        "fecha_alta = EXCLUDED.fecha_alta, "
        "dado_de_alta_por = EXCLUDED.dado_de_alta_por, "
        "rol_al_registrar = EXCLUDED.rol_al_registrar",
        (correo, formatear_nombre_persona(nombre_referencia), area_departamento.strip(), fecha,
         dado_de_alta_por, rol_al_registrar),
    )
    conn.commit()
    conn.close()
    if rol_al_registrar in ("admin", "encargado"):
        etiqueta = ETIQUETAS_ROL.get(rol_al_registrar, rol_al_registrar)
        return True, (f"{correo} quedó autorizado. Al registrarse quedará con rol "
                      f"'{etiqueta}'.")
    return True, f"{correo} quedó autorizado para hacer solicitudes."


def rol_preasignado(correo, db_path: str = None) -> str:
    """
    Rol que debe tomar una cuenta al registrarse, según lo dejó definido el
    admin al autorizar el correo. Devuelve 'admin'/'encargado' si se
    pre-asignó uno, o 'solicitante' por defecto.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT rol_al_registrar FROM correos_autorizados WHERE correo=?",
        ((correo or "").strip().lower(),),
    ).fetchone()
    conn.close()
    if fila and fila[0] in ("admin", "encargado"):
        return fila[0]
    return "solicitante"


def bloquear_correo(correo, db_path: str = None):
    """Deja el correo registrado pero sin permiso (ej. la persona ya no trabaja ahí)."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute(
        "UPDATE correos_autorizados SET estado='bloqueado' WHERE correo=?",
        ((correo or "").strip().lower(),),
    )
    conn.commit()
    conn.close()


def eliminar_correo_autorizado(correo, db_path: str = None):
    """
    Saca el correo de la nómina por completo (a diferencia de bloquear_correo,
    que lo deja registrado con estado 'bloqueado'). No es lo mismo que un
    bloqueo: no queda rastro de que existió, y para volver a autorizarlo hay
    que cargar sus datos de nuevo desde cero.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute(
        "DELETE FROM correos_autorizados WHERE correo=?",
        ((correo or "").strip().lower(),),
    )
    conn.commit()
    conn.close()


def correo_autorizado(correo, db_path: str = None) -> bool:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT estado FROM correos_autorizados WHERE correo=?",
        ((correo or "").strip().lower(),),
    ).fetchone()
    conn.close()
    return bool(fila) and fila[0] == "autorizado"


def listar_correos_autorizados(db_path: str = None) -> pd.DataFrame:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        "SELECT correo, nombre_referencia, area_departamento, estado, fecha_alta "
        "FROM correos_autorizados ORDER BY estado, correo",
        conn,
    )
    conn.close()
    return df


def importar_correos_desde_excel(ruta_archivo, db_path: str = None) -> pd.DataFrame:
    """
    Carga masiva de la nómina de correos municipales. El archivo debe traer
    una columna 'correo' y, opcionalmente, 'nombre' y 'area'.
    """
    db_path = db_path or DB_PATH
    def _es_csv(r):
        return str(getattr(r, "name", r)).lower().endswith((".csv", ".txt"))

    def _leer(header):
        if _es_csv(ruta_archivo):
            return pd.read_csv(ruta_archivo, dtype=str, header=header)
        return pd.read_excel(ruta_archivo, dtype=str, header=header)

    crudo = _leer(None)
    fila_encabezado = None
    for i in range(min(20, len(crudo))):
        valores = [str(v).strip().lower() for v in crudo.iloc[i].tolist()]
        if "correo" in valores:
            fila_encabezado = i
            break
    if fila_encabezado is None:
        raise ValueError("No se encontró una columna llamada 'correo' en el archivo.")

    if hasattr(ruta_archivo, "seek"):
        ruta_archivo.seek(0)
    df = _leer(fila_encabezado)
    df.columns = [str(c).strip().lower() for c in df.columns]

    resultados = []
    for _, fila in df.iterrows():
        correo = fila.get("correo")
        correo = correo.strip() if isinstance(correo, str) else ""
        if not correo:
            continue
        nombre = fila.get("nombre") if isinstance(fila.get("nombre"), str) else ""
        area = fila.get("area") if isinstance(fila.get("area"), str) else ""
        ok, mensaje = autorizar_correo(correo, nombre or "", area or "", "carga masiva", db_path)
        resultados.append({
            "correo": correo, "resultado": "autorizado" if ok else "rechazado", "detalle": mensaje,
        })
    return pd.DataFrame(resultados)


def siguiente_correlativo(db_path: str = None) -> int:
    """
    Número correlativo del formulario físico. Parte en CORRELATIVO_INICIAL
    (2900) para no chocar con el talonario en papel que va en 2798, y avanza
    de uno en uno. Cuando se consigan los datos históricos reales, basta con
    cambiar ese número inicial o cargar los folios antiguos.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    maximo = conn.execute("SELECT MAX(correlativo) FROM solicitudes").fetchone()[0]
    conn.close()
    if maximo is None or maximo < CORRELATIVO_INICIAL:
        return CORRELATIVO_INICIAL
    return int(maximo) + 1


# ---------------------------------------------------- registro de personas

def formato_correo_valido(correo: str, dominios_permitidos=None) -> bool:
    """
    Valida solo que el correo tenga forma de correo (algo@algo.algo).

    Ya NO se exige un dominio institucional determinado: el control real es
    la nómina de correos que autoriza el encargado. Restringir además por
    dominio dejaba fuera casos legítimos (personal a honorarios, direcciones
    con dominio propio, convenios con otros servicios) sin agregar seguridad,
    porque cualquiera puede inventar una dirección del dominio correcto.

    El parámetro dominios_permitidos se mantiene por compatibilidad; si se
    entrega una lista, se sigue exigiendo, pero por defecto no se usa.
    """
    correo = (correo or "").strip().lower()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo):
        return False
    if dominios_permitidos:
        dominio = correo.split("@")[-1]
        return dominio in [d.strip().lower() for d in dominios_permitidos]
    return True


def _hash_password(password: str, salt: bytes = None):
    """
    Guarda la contraseña como hash, nunca en texto plano: si alguien abre el
    archivo bodega.db no puede leer las contraseñas de nadie. Se usa PBKDF2
    con SHA-256, que viene en la librería estándar de Python (sin instalar
    nada extra) y es el mecanismo recomendado para este caso.
    """
    if salt is None:
        salt = os.urandom(16)
    hash_bytes = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120_000)
    return hash_bytes.hex(), salt.hex()


def validar_password(password: str, confirmacion: str):
    """Reglas mínimas de contraseña. Devuelve (ok, mensaje)."""
    password = password or ""
    if len(password) < 6:
        return False, "La contraseña debe tener al menos 6 caracteres."
    if password != (confirmacion or ""):
        return False, "Las dos contraseñas no coinciden. Vuelva a escribirlas."
    return True, ""


def registrar_persona(correo, nombre, area_departamento, nombre_supervisor, password,
                      db_path: str = None, rol="solicitante"):
    """
    Registra a la persona con su contraseña. Ya no se pide el correo del
    supervisor: basta su nombre, que es lo que va impreso en el formulario
    para la firma.
    """
    db_path = db_path or DB_PATH
    correo = correo.strip().lower()
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")
    pass_hash, salt = _hash_password(password)
    conn = get_connection(db_path)
    conn.execute(
        "INSERT INTO personas_registradas "
        "(correo, nombre, area_departamento, nombre_supervisor, correo_supervisor, "
        " password_hash, password_salt, fecha_registro, rol) "
        "VALUES (?, ?, ?, ?, '', ?, ?, ?, ?) "
        "ON CONFLICT (correo) DO UPDATE SET "
        "nombre = EXCLUDED.nombre, "
        "area_departamento = EXCLUDED.area_departamento, "
        "nombre_supervisor = EXCLUDED.nombre_supervisor, "
        "correo_supervisor = EXCLUDED.correo_supervisor, "
        "password_hash = EXCLUDED.password_hash, "
        "password_salt = EXCLUDED.password_salt, "
        "fecha_registro = EXCLUDED.fecha_registro, "
        "rol = EXCLUDED.rol",
        (correo, formatear_nombre_persona(nombre), area_departamento.strip(),
         formatear_nombre_persona(nombre_supervisor), pass_hash, salt, fecha, rol),
    )
    conn.commit()
    conn.close()


# Cuenta de encargado que se crea sola la primera vez, para que exista un
# acceso inicial sin depender de nadie más.
#
# La contraseña YA NO está escrita acá. Se lee de la configuración del
# entorno, en este orden:
#   1. st.secrets["BODEGA_PASS_ENCARGADO"]  (Streamlit Cloud, o el archivo
#      local .streamlit/secrets.toml que no se sube al repositorio)
#   2. la variable de entorno BODEGA_PASS_ENCARGADO
# Si no hay ninguna de las dos, la cuenta no se crea y la aplicación lo avisa
# en pantalla, en vez de quedar con una clave conocida por cualquiera que
# haya visto el código.

ENCARGADO_POR_DEFECTO = {
    "correo": os.environ.get("BODEGA_CORREO_ENCARGADO", "g.fierro03@ufromail.cl"),
    "nombre": os.environ.get("BODEGA_NOMBRE_ENCARGADO", "Gonzalo Fierro Cea"),
    "area": "Bodega Municipal",
}


def _password_encargado():
    """Busca la contraseña inicial en los secretos o en el entorno."""
    try:
        import streamlit as st
        if "BODEGA_PASS_ENCARGADO" in st.secrets:
            return str(st.secrets["BODEGA_PASS_ENCARGADO"])
    except Exception:
        # fuera de Streamlit (scripts, pruebas) no hay secrets: se sigue al entorno
        pass
    return os.environ.get("BODEGA_PASS_ENCARGADO")


def asegurar_encargado_por_defecto(db_path: str = None):
    """
    Crea (una sola vez) la cuenta de encargado y la deja autorizada.
    Devuelve (ok, mensaje): si no hay contraseña configurada, no crea nada y
    explica qué falta.
    """
    db_path = db_path or DB_PATH
    correo = ENCARGADO_POR_DEFECTO["correo"]
    autorizar_correo(correo, ENCARGADO_POR_DEFECTO["nombre"],
                     ENCARGADO_POR_DEFECTO["area"], "sistema", db_path)

    persona = obtener_persona(correo, db_path)
    if persona is not None:
        # La cuenta del creador es SIEMPRE admin: si una versión anterior la
        # dejó como 'encargado', se asciende acá (idempotente). Así el creador
        # nunca pierde la capacidad de gestionar roles.
        if persona.get("rol") != "admin":
            conn = get_connection(db_path)
            conn.execute("UPDATE personas_registradas SET rol='admin' WHERE correo=?", (correo,))
            conn.commit()
            conn.close()
        return True, ""

    password = _password_encargado()
    if not password:
        return False, (
            "No hay contraseña de encargado configurada, así que la cuenta inicial no se creó. "
            "Defina BODEGA_PASS_ENCARGADO en los secretos de Streamlit (o como variable de "
            "entorno) y recargue."
        )

    registrar_persona(
        correo, ENCARGADO_POR_DEFECTO["nombre"], ENCARGADO_POR_DEFECTO["area"],
        "", password, db_path, rol="admin",
    )
    guardar_config("nombre_encargado", ENCARGADO_POR_DEFECTO["nombre"], db_path)
    return True, ""


# Roles del sistema, de mayor a menor:
#   'admin'      -> usa la app de encargado Y puede dar/quitar roles (crear
#                   encargados y otros admin). Son el creador y el supervisor.
#   'encargado'  -> usa la app de encargado (procesar solicitudes, inventario,
#                   etc.) pero NO puede gestionar roles.
#   'solicitante'-> hace pedidos desde la web.
ROLES_CON_ACCESO_ENCARGADO = ("admin", "encargado")


def es_encargado(correo, db_path: str = None) -> bool:
    """True si la cuenta puede entrar a la app de encargado (admin o encargado)."""
    db_path = db_path or DB_PATH
    persona = obtener_persona(correo, db_path)
    return bool(persona) and persona.get("rol") in ROLES_CON_ACCESO_ENCARGADO


def es_admin(correo, db_path: str = None) -> bool:
    """True solo para los admin (creador y supervisor): los únicos que pueden
    dar/quitar el rol de encargado a otras cuentas."""
    db_path = db_path or DB_PATH
    persona = obtener_persona(correo, db_path)
    return bool(persona) and persona.get("rol") == "admin"


def cambiar_password(correo, password_actual, password_nuevo, confirmacion,
                     db_path: str = None):
    """
    Cambia la contraseña de una cuenta. Pide la actual para que nadie pueda
    cambiarla desde una sesión ajena que quedó abierta.
    """
    db_path = db_path or DB_PATH
    ok, _ = verificar_login(correo, password_actual, db_path)
    if not ok:
        return False, "La contraseña actual no es correcta."
    ok, mensaje = validar_password(password_nuevo, confirmacion)
    if not ok:
        return False, mensaje

    pass_hash, salt = _hash_password(password_nuevo)
    conn = get_connection(db_path)
    conn.execute(
        "UPDATE personas_registradas SET password_hash=?, password_salt=? WHERE correo=?",
        (pass_hash, salt, correo.strip().lower()),
    )
    conn.commit()
    conn.close()
    return True, "Contraseña actualizada."


def verificar_login(correo, password, db_path: str = None):
    """
    Devuelve (ok, mensaje). Solo entra quien esté registrado, tenga la
    contraseña correcta y siga autorizado en la nómina (si al encargado le
    bloquean el correo, deja de poder entrar aunque sepa su contraseña).
    """
    db_path = db_path or DB_PATH
    correo = (correo or "").strip().lower()
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT password_hash, password_salt FROM personas_registradas WHERE correo = ?",
        (correo,),
    ).fetchone()
    conn.close()

    if fila is None:
        return False, "Ese correo no está registrado todavía. Regístrese primero."
    if not fila[0] or not fila[1]:
        return False, ("Esta cuenta se creó antes de que existieran las contraseñas. "
                       "Pida al encargado que la elimine para volver a registrarse.")
    if not correo_autorizado(correo, db_path):
        return False, "Este correo ya no está autorizado para hacer solicitudes."

    intento, _ = _hash_password(password or "", bytes.fromhex(fila[1]))
    if intento != fila[0]:
        return False, "Contraseña incorrecta."
    return True, ""


def obtener_persona(correo, db_path: str = None):
    """Devuelve dict con los datos de la persona registrada, o None si no está registrada."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT nombre, area_departamento, nombre_supervisor, correo_supervisor, fecha_registro, "
        "       COALESCE(rol, 'solicitante') "
        "FROM personas_registradas WHERE correo = ?",
        (correo.strip().lower(),),
    ).fetchone()
    conn.close()
    if fila is None:
        return None
    return {
        "nombre": fila[0], "area_departamento": fila[1],
        "nombre_supervisor": fila[2] or "", "correo_supervisor": fila[3],
        "fecha_registro": fila[4], "rol": fila[5],
    }


def actualizar_datos_persona(correo, nombre, area_departamento, nombre_supervisor,
                             db_path: str = None):
    """
    Actualiza los datos de la cuenta (nombre, área/departamento, supervisor).
    NO toca la contraseña ni el rol. La oficina no se guarda acá: se pide en
    cada solicitud. Devuelve (ok, mensaje).
    """
    db_path = db_path or DB_PATH
    correo = (correo or "").strip().lower()
    nombre = (nombre or "").strip()
    area = (area_departamento or "").strip()
    supervisor = (nombre_supervisor or "").strip()
    if not nombre:
        return False, "El nombre no puede quedar vacío."
    if not area:
        return False, "El área / departamento no puede quedar vacío."
    conn = get_connection(db_path)
    existe = conn.execute(
        "SELECT 1 FROM personas_registradas WHERE correo=?", (correo,)
    ).fetchone()
    if existe is None:
        conn.close()
        return False, "La cuenta no existe."
    conn.execute(
        "UPDATE personas_registradas SET nombre=?, area_departamento=?, nombre_supervisor=? "
        "WHERE correo=?",
        (nombre, area, supervisor, correo),
    )
    conn.commit()
    conn.close()
    return True, "Datos actualizados."


# ---------------------------------------------- gestión de encargados (roles)

def listar_personas_registradas(db_path: str = None) -> pd.DataFrame:
    """Personas registradas con su rol, para poder gestionar quién es
    encargado. Encargados primero."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        "SELECT correo, nombre, area_departamento, COALESCE(rol, 'solicitante') AS rol "
        "FROM personas_registradas ORDER BY (COALESCE(rol,'solicitante')='encargado') DESC, nombre",
        conn,
    )
    conn.close()
    return df


def contar_admins(db_path: str = None) -> int:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    n = conn.execute(
        "SELECT COUNT(*) FROM personas_registradas WHERE COALESCE(rol,'solicitante')='admin'"
    ).fetchone()[0]
    conn.close()
    return int(n)


ETIQUETAS_ROL = {"admin": "Administrador", "encargado": "Encargado", "solicitante": "Solicitante"}


def cambiar_rol(correo, nuevo_rol, db_path: str = None):
    """
    Cambia el rol de una cuenta ya registrada (solo la ejecutan los admin,
    ver es_admin). Reglas:
      - roles válidos: 'admin', 'encargado', 'solicitante'.
      - para dar acceso de encargado o admin, el correo tiene que estar en la
        nómina autorizada (si no, primero hay que autorizarlo).
      - un administrador NO le puede quitar/cambiar el rol a otro
        administrador: los admin son pares y no se degradan entre sí desde la
        interfaz (a un encargado sí se le puede cambiar). Si alguna vez hiciera
        falta sacar un admin, se hace directo en la base de datos.
    Devuelve (ok, mensaje).
    """
    db_path = db_path or DB_PATH
    correo = (correo or "").strip().lower()
    if nuevo_rol not in ("admin", "encargado", "solicitante"):
        return False, "Rol inválido."

    persona = obtener_persona(correo, db_path)
    if persona is None:
        return False, "Esa cuenta no está registrada (la persona debe registrarse primero)."
    rol_actual = persona.get("rol")
    if rol_actual == nuevo_rol:
        return False, f"La cuenta ya es {ETIQUETAS_ROL.get(nuevo_rol, nuevo_rol)}."

    # Los administradores no se tocan entre sí: una cuenta que ya es admin no
    # puede ser modificada desde la interfaz (ni degradada ni cambiada).
    if rol_actual == "admin":
        return False, ("No se puede cambiar el rol de un administrador. Entre administradores "
                       "no se quitan permisos.")

    if nuevo_rol in ROLES_CON_ACCESO_ENCARGADO and not correo_autorizado(correo, db_path):
        return False, ("Ese correo no está en la nómina autorizada. Autorícelo primero "
                       "en 'Autorizar un correo' y después asígnele el rol.")

    conn = get_connection(db_path)
    conn.execute("UPDATE personas_registradas SET rol=? WHERE correo=?", (nuevo_rol, correo))
    conn.commit()
    conn.close()
    return True, f"{correo} ahora es {ETIQUETAS_ROL.get(nuevo_rol, nuevo_rol)}."


# ---------------------------------------------------- flujo de la solicitud

def crear_solicitud(solicitante, supervisor, area_departamento, items, db_path: str = None,
                     correo_solicitante=None, correo_supervisor=None, oficina=None):
    """
    items: lista de (codigo_producto, cantidad_solicitada).
    Estado inicial: 'pendiente_firma' (a la espera de que el solicitante
    vuelva con el papel timbrado y firmado).

    VALIDACIÓN DURA: si falta solicitante, supervisor, área, o no hay
    productos, NO se guarda absolutamente nada — se lanza ValueError antes
    de tocar la base. Una solicitud incompleta nunca llega a existir como
    registro, así que no hay nada que "eliminar" después: nunca se creó.
    """
    db_path = db_path or DB_PATH
    faltantes = []
    if not (solicitante or "").strip():
        faltantes.append("solicitante")
    if not (supervisor or "").strip():
        faltantes.append("supervisor")
    if not (area_departamento or "").strip():
        faltantes.append("área/departamento")
    if not items:
        faltantes.append("productos (la lista está vacía)")
    if faltantes:
        raise ValueError(
            "No se puede registrar la solicitud — faltan datos obligatorios: " + ", ".join(faltantes)
        )

    # Departamento y oficina se guardan SIEMPRE en MAYÚSCULAS, el mismo
    # formato que ya usa el comprobante impreso (ver datos_para_impresion).
    # Así "DIDECO", "dideco" y "Dideco" quedan como un solo valor y no se
    # parten en categorías distintas al graficar las estadísticas.
    area_departamento = (area_departamento or "").strip().upper()
    oficina = (oficina or "").strip().upper()
    # Nombres de personas con formato uniforme (inicial en mayúscula, resto en
    # minúscula) para que TODAS las solicitudes queden igual, sin importar si el
    # nombre vino del registro o se tipeó a mano al crear/editar el pedido. Ver
    # formatear_nombre_persona: respeta apellidos compuestos ("San Miguel").
    solicitante = formatear_nombre_persona(solicitante)
    supervisor = formatear_nombre_persona(supervisor)

    correlativo = siguiente_correlativo(db_path)
    folio = f"Solicitud-{correlativo}"
    fecha = ahora_chile().strftime("%Y-%m-%d %H:%M")
    conn = get_connection(db_path)
    cur = conn.execute(
        "INSERT INTO solicitudes "
        "(folio, fecha_solicitud, solicitante, supervisor, area_departamento, estado, "
        " correo_solicitante, correo_supervisor, correlativo, oficina) "
        "VALUES (?, ?, ?, ?, ?, 'pendiente_firma', ?, ?, ?, ?) "
        "RETURNING id",
        (folio, fecha, solicitante, supervisor, area_departamento, correo_solicitante,
         correo_supervisor, correlativo, oficina),
    )
    solicitud_id = cur.lastrowid

    for codigo, cantidad in items:
        cantidad_i = formatear_cantidad(cantidad)
        estado_val, mensaje = validar_disponibilidad(codigo, cantidad_i, db_path)
        conn.execute(
            "INSERT INTO solicitud_detalle "
            "(solicitud_id, codigo_producto, cantidad_solicitada, cantidad_entregada, mensaje_sistema) "
            "VALUES (?, ?, ?, NULL, ?)",
            (solicitud_id, codigo, cantidad_i, mensaje),
        )
    conn.commit()
    conn.close()
    return folio


def aceptar_preliminar(folio, db_path: str = None):
    """El solicitante volvió con el papel timbrado/firmado; el encargado
    marca 'aceptado con posibilidad a cambio' y va a la bodega a verificar/entregar."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET estado='preliminar_aceptada' WHERE folio=?", (folio,))
    conn.commit()
    conn.close()


def editar_entrega(folio, codigo_producto, cantidad_entregada, db_path: str = None):
    """El encargado ajusta lo realmente entregado (puede diferir de lo solicitado
    si en bodega hay menos de lo que decía el sistema)."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    solicitud_id = conn.execute(
        "SELECT id FROM solicitudes WHERE folio=?", (folio,)
    ).fetchone()[0]
    conn.execute(
        "UPDATE solicitud_detalle SET cantidad_entregada=? "
        "WHERE solicitud_id=? AND codigo_producto=?",
        (formatear_cantidad(cantidad_entregada), solicitud_id, codigo_producto),
    )
    conn.execute("UPDATE solicitudes SET estado='editada' WHERE folio=?", (folio,))
    conn.commit()
    conn.close()


# Prefijo de código para productos creados a mano al procesar un pedido: algo
# que el solicitante trajo escrito y timbrado pero que no estaba en el catálogo
# SMC. Sirve para distinguirlos — a estos NO se les descuenta stock al cerrar,
# porque no hay un saldo real registrado (ver cerrar_solicitud).
PREFIJO_PRODUCTO_MANUAL = "MAN-"


def es_producto_manual(codigo) -> bool:
    return str(codigo).startswith(PREFIJO_PRODUCTO_MANUAL)


def crear_producto_manual(nombre, unidad_medida="", valor_unitario=0,
                          categoria="SIN CATEGORÍA", db_path: str = None) -> str:
    """
    Crea un producto que no estaba en el catálogo, para poder sumarlo a un
    pedido cuando el solicitante trajo algo escrito a mano y timbrado. Le asigna
    un código propio con prefijo MAN- (no choca con los códigos numéricos del
    catálogo SMC), lo deja activo y buscable (registra su nombre como alias) y
    arranca sin stock (saldo 0). Devuelve el código generado.
    """
    db_path = db_path or DB_PATH
    nombre = (nombre or "").strip()
    if not nombre:
        raise ValueError("El producto nuevo necesita un nombre.")
    conn = get_connection(db_path)
    # Se evita duplicar un nombre que ya existe en el catálogo: en ese caso hay
    # que buscarlo y agregarlo, no crear un gemelo (además, varias rutas ubican
    # el producto por nombre_estandar y un nombre repetido las volvería ambiguas).
    existe = conn.execute(
        "SELECT codigo FROM productos WHERE nombre_estandar=?", (nombre,)
    ).fetchone()
    if existe:
        conn.close()
        raise ValueError(
            f'Ya existe un producto llamado "{nombre}" en el catálogo: '
            "búsquelo con el buscador y agréguelo, en vez de crear uno nuevo."
        )
    # Siguiente número de la secuencia MAN-0001, MAN-0002, ...
    codigos = conn.execute(
        "SELECT codigo FROM productos WHERE codigo LIKE 'MAN-%'"
    ).fetchall()
    maximo = 0
    for (c,) in codigos:
        try:
            maximo = max(maximo, int(str(c).split("-", 1)[1]))
        except (IndexError, ValueError):
            pass
    codigo = f"{PREFIJO_PRODUCTO_MANUAL}{maximo + 1:04d}"
    conn.execute(
        "INSERT INTO productos "
        "(codigo, nombre_estandar, unidad_medida, categoria, saldo, stock_critico, "
        " valor_saldo, valor_unitario, activo) "
        "VALUES (?, ?, ?, ?, 0, 0, 0, ?, 1)",
        (codigo, nombre, (unidad_medida or "").strip(),
         (categoria or "SIN CATEGORÍA").strip(), float(valor_unitario or 0)),
    )
    # el nombre también queda como su propio alias de búsqueda (igual que en cargar_catalogo)
    conn.execute(
        "INSERT INTO alias_productos (texto_alias, texto_alias_normalizado, codigo_producto) "
        "VALUES (?, ?, ?)",
        (nombre, normalizar(nombre), codigo),
    )
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()
    return codigo


def agregar_producto_a_solicitud(folio, codigo_producto, cantidad, db_path: str = None):
    """
    Suma una línea de producto a un pedido en curso (no cerrado ni anulado): el
    caso típico es algo escrito a mano y timbrado que no venía en el pedido
    digital. Deja el pedido en estado 'editada' y evita duplicar un producto que
    ya estuviera en el pedido.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute(
        "SELECT id, estado FROM solicitudes WHERE folio=?", (folio,)
    ).fetchone()
    if fila is None:
        conn.close()
        raise ValueError("Folio no encontrado.")
    solicitud_id, estado = fila
    if estado in ("cerrada", "anulada"):
        conn.close()
        raise ValueError(f"No se puede agregar productos a una solicitud '{estado}'.")
    ya_esta = conn.execute(
        "SELECT 1 FROM solicitud_detalle WHERE solicitud_id=? AND codigo_producto=?",
        (solicitud_id, codigo_producto),
    ).fetchone()
    if ya_esta:
        conn.close()
        raise ValueError("Ese producto ya está en el pedido; ajuste su cantidad en la lista.")
    cantidad_i = formatear_cantidad(cantidad)
    if es_producto_manual(codigo_producto):
        mensaje = "Producto agregado a mano (fuera de catálogo)."
    else:
        _, mensaje = validar_disponibilidad(codigo_producto, cantidad_i, db_path)
    conn.execute(
        "INSERT INTO solicitud_detalle "
        "(solicitud_id, codigo_producto, cantidad_solicitada, cantidad_entregada, mensaje_sistema) "
        "VALUES (?, ?, ?, NULL, ?)",
        (solicitud_id, codigo_producto, cantidad_i, mensaje),
    )
    conn.execute("UPDATE solicitudes SET estado='editada' WHERE folio=?", (folio,))
    conn.commit()
    conn.close()


def modificar_cantidad_solicitada(folio, codigo_producto, nueva_cantidad, db_path: str = None):
    """
    Potestad del encargado: corregir lo que el solicitante pidió (ej. el
    solicitante se equivocó de cantidad, o pidió algo que ya no corresponde),
    ANTES de que la solicitud se cierre. Re-valida disponibilidad contra el
    nuevo número.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    solicitud_id, estado = conn.execute(
        "SELECT id, estado FROM solicitudes WHERE folio=?", (folio,)
    ).fetchone()
    if estado in ("cerrada", "anulada"):
        conn.close()
        raise ValueError(f"No se puede modificar una solicitud en estado '{estado}'.")

    cantidad_i = formatear_cantidad(nueva_cantidad)
    _, mensaje = validar_disponibilidad(codigo_producto, cantidad_i, db_path)
    conn.execute(
        "UPDATE solicitud_detalle SET cantidad_solicitada=?, mensaje_sistema=? "
        "WHERE solicitud_id=? AND codigo_producto=?",
        (cantidad_i, mensaje, solicitud_id, codigo_producto),
    )
    conn.execute("UPDATE solicitudes SET estado='editada' WHERE folio=?", (folio,))
    conn.commit()
    conn.close()


def anular_solicitud(folio, motivo, db_path: str = None):
    """
    Potestad del encargado: anular una solicitud completa antes de cerrarla
    (ej. el solicitante nunca volvió con el papel firmado, se duplicó, o se
    canceló la necesidad). No descuenta stock porque una solicitud anulada
    nunca llegó a 'cerrada'. Queda con motivo registrado para trazabilidad.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    estado_actual = conn.execute(
        "SELECT estado FROM solicitudes WHERE folio=?", (folio,)
    ).fetchone()
    if estado_actual is None:
        conn.close()
        raise ValueError("Folio no encontrado.")
    if estado_actual[0] == "cerrada":
        conn.close()
        raise ValueError("No se puede anular una solicitud ya cerrada (el stock ya se descontó).")
    conn.execute(
        "UPDATE solicitudes SET estado='anulada', motivo_anulacion=? WHERE folio=?",
        (motivo, folio),
    )
    conn.commit()
    conn.close()


def cerrar_solicitud(folio, db_path: str = None, usuario_operacion=None):
    """Descuenta stock real según cantidad_entregada, cierra la solicitud
    y evalúa alertas de stock agotado / bajo stock crítico.

    usuario_operacion: quién realizó el movimiento en bodega. Queda guardado
    porque el comprobante impreso lo lleva ('USUARIO' del formulario)."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    solicitud_id = conn.execute(
        "SELECT id FROM solicitudes WHERE folio=?", (folio,)
    ).fetchone()[0]
    detalle = conn.execute(
        "SELECT d.codigo_producto, d.cantidad_solicitada, d.cantidad_entregada, "
        "       COALESCE(p.valor_unitario, 0) "
        "FROM solicitud_detalle d JOIN productos p ON p.codigo = d.codigo_producto "
        "WHERE d.solicitud_id=?",
        (solicitud_id,),
    ).fetchall()

    alertas = []
    for codigo, cant_sol, cant_ent, unitario in detalle:
        cantidad_real = formatear_cantidad(cant_ent if cant_ent is not None else cant_sol)
        valor_movimiento = float(unitario or 0) * cantidad_real
        # Se descuenta el stock Y el valor en pesos a la vez, manteniendo el
        # invariante valor_saldo = valor_unitario * saldo: así el valor del
        # inventario baja solo con cada entrega, usando el valor unitario fijo
        # del corte (no uno recalculado). El '(saldo - ?)' del SET usa el saldo
        # ANTERIOR a este UPDATE (así lo evalúan tanto SQLite como Postgres),
        # por lo que queda valor_unitario * (saldo nuevo).
        # Los productos agregados a mano (MAN-) no tienen un saldo real
        # registrado: no se les descuenta stock (dejaría el inventario en
        # negativo con una alerta falsa de 'agotado'). Igual se deja constancia
        # de la entrega en solicitud_detalle (cantidad y valor), más abajo.
        if not es_producto_manual(codigo):
            conn.execute(
                "UPDATE productos SET saldo = saldo - ?, "
                "valor_saldo = valor_unitario * (saldo - ?) WHERE codigo = ?",
                (cantidad_real, cantidad_real, codigo),
            )
        conn.execute(
            "UPDATE solicitud_detalle SET cantidad_entregada=?, valor_movimiento=? "
            "WHERE solicitud_id=? AND codigo_producto=?",
            (cantidad_real, valor_movimiento, solicitud_id, codigo),
        )
    conn.commit()
    conn.close()

    for fila in detalle:
        codigo = fila[0]
        if es_producto_manual(codigo):
            continue  # sin saldo real: no corresponde alertar por stock
        tipo, mensaje = evaluar_alerta_stock(codigo, db_path)
        if tipo != "ok":
            alertas.append((codigo, tipo, mensaje))

    conn = get_connection(db_path)
    conn.execute(
        "UPDATE solicitudes SET estado='cerrada', usuario_operacion=? WHERE folio=?",
        (usuario_operacion or nombre_encargado(db_path), folio),
    )
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()  # se descontó stock: refrescar inventario cacheado
    return alertas


# --------------------------------------------------------------- reportería

def resumen_solicitud(folio, db_path: str = None) -> pd.DataFrame:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.folio, s.estado, s.fecha_solicitud, s.solicitante, s.supervisor, s.area_departamento,
               p.nombre_estandar AS producto, d.cantidad_solicitada, d.cantidad_entregada, d.mensaje_sistema
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        WHERE s.folio = ?
        """,
        conn,
        params=(folio,),
    )
    conn.close()
    return _normalizar_columna_mensaje(df)


def listar_solicitudes_activas(db_path: str = None, incluir_anteriores=False) -> pd.DataFrame:
    """
    Pantalla de trabajo del encargado. Muestra:

    Muestra únicamente las solicitudes PENDIENTES, sin importar de qué día
    sean: nunca se ocultan, porque si algo quedó sin entregar tiene que
    seguir a la vista. Las que vienen de jornadas anteriores se marcan como
    atrasadas.

    Las cerradas y anuladas no aparecen acá: al terminar el proceso salen de
    la pantalla de trabajo y quedan disponibles en el Historial y en Pedidos
    completados.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.folio, s.correlativo, s.estado, s.fecha_solicitud, s.solicitante,
               s.area_departamento, s.oficina, COUNT(d.id) AS n_productos
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        WHERE s.estado NOT IN ('cerrada', 'anulada')
        GROUP BY s.id
        ORDER BY s.fecha_solicitud DESC
        """,
        conn,
    )
    conn.close()
    if df.empty:
        return df

    corte = inicio_jornada_actual()
    fechas = pd.to_datetime(df["fecha_solicitud"], errors="coerce")

    # atrasada = viene de una jornada anterior y sigue sin cerrarse
    df["atrasada"] = fechas < corte
    return df


def contar_solicitudes_atrasadas(db_path: str = None) -> int:
    """Pendientes que vienen arrastradas de jornadas anteriores."""
    db_path = db_path or DB_PATH
    df = listar_solicitudes_activas(db_path)
    if df.empty:
        return 0
    return int(df["atrasada"].sum())


def listar_inventario_general(db_path: str = None) -> pd.DataFrame:
    """
    Inventario general para el encargado. No incluye stock_critico a propósito
    (esa columna ya se destaca en la pestaña Inventario Crítico, no hace
    falta repetirla acá). El valor de cada línea es el 'SALDO VAL.' tal como
    se importó del último corte de SMC — no se deriva ni recalcula ningún
    precio unitario, porque no existe uno confiable.
    """
    db_path = db_path or DB_PATH

    def _cargar():
        conn = get_connection(db_path)
        df = pd.read_sql(
            """
            SELECT codigo, nombre_estandar, categoria, unidad_medida, saldo, valor_saldo
            FROM productos
            WHERE activo = 1
            ORDER BY categoria, nombre_estandar
            """,
            conn,
        )
        conn.close()
        df["saldo"] = df["saldo"].apply(formatear_cantidad)
        df["valor_saldo"] = df["valor_saldo"].round(0).astype(int)
        return df

    return _leer_cacheado(("inventario_general", db_path), _cargar)


def valor_total_inventario(db_path: str = None) -> int:
    """Valor total de los bienes en bodega: suma de 'SALDO VAL.' del último corte importado."""
    db_path = db_path or DB_PATH

    def _cargar():
        conn = get_connection(db_path)
        # Se suma en Python (doble precisión) en vez de SUM() en la base: en
        # Postgres valor_saldo es REAL (precisión simple) y SUM() sobre un
        # total de decenas de millones pierde algunos pesos. Cada valor por
        # producto sí es exacto, así que sumarlos en Python da el total exacto.
        valores = conn.execute(
            "SELECT valor_saldo FROM productos WHERE activo = 1"
        ).fetchall()
        conn.close()
        total = sum(float(v[0]) for v in valores if v[0] is not None)
        return int(round(total))

    return _leer_cacheado(("valor_total_inventario", db_path), _cargar)


def listar_stock_critico(db_path: str = None) -> pd.DataFrame:
    """Insumos agotados o bajo su stock crítico — los que requieren compra
    o renovación más urgente."""
    db_path = db_path or DB_PATH

    def _cargar():
        conn = get_connection(db_path)
        df = pd.read_sql(
            """
            SELECT codigo, nombre_estandar, categoria, unidad_medida, saldo, stock_critico,
                   CASE WHEN saldo <= 0 THEN 'AGOTADO' ELSE 'BAJO CRÍTICO' END AS urgencia
            FROM productos
            WHERE activo = 1 AND (saldo <= 0 OR (stock_critico > 0 AND saldo < stock_critico))
            ORDER BY (saldo <= 0) DESC, saldo ASC
            """,
            conn,
        )
        conn.close()
        for col in ("saldo", "stock_critico"):
            df[col] = df[col].apply(formatear_cantidad)
        return df

    return _leer_cacheado(("stock_critico", db_path), _cargar)


def historial_periodos(db_path: str = None):
    """Lista los periodos AÑO-MES (ej. '2026-07') que tienen solicitudes,
    del más reciente al más antiguo — para armar los sub-índices del historial."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql("SELECT fecha_solicitud FROM solicitudes", conn)
    conn.close()
    if df.empty:
        return []
    periodos = sorted({f[:7] for f in df["fecha_solicitud"]}, reverse=True)
    return periodos


def historial_por_periodo(periodo: str, db_path: str = None) -> pd.DataFrame:
    """periodo en formato 'YYYY-MM'."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.folio, s.fecha_solicitud, s.estado, s.solicitante, s.area_departamento,
               p.codigo, p.nombre_estandar AS producto,
               d.cantidad_solicitada, d.cantidad_entregada, d.mensaje_sistema
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        WHERE s.fecha_solicitud LIKE ?
        ORDER BY s.fecha_solicitud DESC
        """,
        conn,
        params=(f"{periodo}%",),
    )
    conn.close()
    for col in ("cantidad_solicitada", "cantidad_entregada"):
        df[col] = df[col].apply(formatear_cantidad)
    return _normalizar_columna_mensaje(df)


HORA_CIERRE_JORNADA = 19  # a las 19:00 se cierra la jornada de bodega


def inicio_jornada_actual(ahora=None):
    """
    Devuelve el instante en que empezó la jornada vigente. La jornada corre
    de 19:00 a 19:00: pasada esa hora, lo del día anterior deja de aparecer
    en la pantalla de solicitudes activas (pero sigue en el historial).
    """
    ahora = ahora or ahora_chile()
    corte_hoy = ahora.replace(hour=HORA_CIERRE_JORNADA, minute=0, second=0, microsecond=0)
    if ahora >= corte_hoy:
        return corte_hoy
    return corte_hoy - timedelta(days=1)


def pedidos_completados_jornada(db_path: str = None) -> pd.DataFrame:
    """
    Solicitudes cerradas de la jornada vigente (desde el último corte de las
    19:00). Pasada esa hora, lo cerrado el día anterior deja de aparecer en
    "Pedidos completados" — sigue disponible siempre en el Historial, que no
    tiene corte.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        "SELECT folio, correlativo, fecha_solicitud, solicitante, area_departamento "
        "FROM solicitudes WHERE estado='cerrada' ORDER BY correlativo DESC",
        conn,
    )
    conn.close()
    if df.empty:
        return df
    corte = inicio_jornada_actual()
    fechas = pd.to_datetime(df["fecha_solicitud"], errors="coerce")
    return df[fechas >= corte].reset_index(drop=True)


def _filtro_fechas_sql(desde, hasta):
    """Arma la condición WHERE de fechas y sus parámetros."""
    condiciones, params = [], []
    if desde is not None:
        condiciones.append("s.fecha_solicitud >= ?")
        params.append(f"{desde} 00:00")
    if hasta is not None:
        condiciones.append("s.fecha_solicitud <= ?")
        params.append(f"{hasta} 23:59")
    return (" AND " + " AND ".join(condiciones)) if condiciones else "", params


def estadisticas_consumo(desde=None, hasta=None, solo_cerradas=True, db_path: str = None):
    """
    Devuelve un diccionario de DataFrames con el consumo agregado, para las
    gráficas del panel de estadísticas.

    La agregación se hace en SQL y no en pandas: con miles de solicitudes
    históricas, traerlas todas a memoria para contarlas sería lento sin
    necesidad.

    Se usa la cantidad entregada cuando existe (lo que realmente salió de
    bodega) y, si no, la solicitada. Además de unidades y solicitudes se
    calcula el VALOR en pesos de lo consumido: se usa valor_movimiento (el
    valor fijado al cerrar cada línea) y, para comprobantes cerrados antes de
    que existiera esa columna, se estima con valor_unitario * cantidad.

    Departamento y oficina se agrupan en MAYÚSCULAS (UPPER) para que las
    variantes de tipeo del mismo nombre —"DIDECO", "dideco", "Dideco"— caigan
    en un solo grupo. Las solicitudes nuevas ya se guardan en mayúsculas (ver
    crear_solicitud); el UPPER acá cubre además los datos históricos mixtos.
    """
    db_path = db_path or DB_PATH
    filtro_fecha, params = _filtro_fechas_sql(desde, hasta)
    filtro_estado = "s.estado = 'cerrada'" if solo_cerradas else "s.estado != 'anulada'"
    base_where = f"WHERE {filtro_estado}{filtro_fecha}"
    cantidad = "COALESCE(d.cantidad_entregada, d.cantidad_solicitada)"
    # Valor en pesos de cada línea: el fijado al cerrar, o estimado si es viejo.
    valor = ("COALESCE(d.valor_movimiento, "
             "COALESCE(p.valor_unitario, 0) * COALESCE(d.cantidad_entregada, d.cantidad_solicitada))")

    conn = get_connection(db_path)

    def _consulta(campo, alias, mayus=False):
        etiqueta = f"COALESCE(NULLIF(TRIM({campo}), ''), '(sin dato)')"
        if mayus:
            etiqueta = f"UPPER({etiqueta})"
        return pd.read_sql(
            f"""
            SELECT {etiqueta} AS {alias},
                   COUNT(DISTINCT s.id) AS solicitudes,
                   SUM({cantidad}) AS unidades,
                   SUM({valor}) AS valor
            FROM solicitudes s
            JOIN solicitud_detalle d ON d.solicitud_id = s.id
            JOIN productos p ON p.codigo = d.codigo_producto
            {base_where}
            GROUP BY {alias}
            ORDER BY unidades DESC
            """,
            conn, params=params,
        )

    por_departamento = _consulta("s.area_departamento", "departamento", mayus=True)
    por_oficina = _consulta("s.oficina", "oficina", mayus=True)
    por_solicitante = _consulta("s.solicitante", "solicitante")

    por_producto = pd.read_sql(
        f"""
        SELECT p.nombre_estandar AS producto, p.categoria,
               COUNT(DISTINCT s.id) AS veces_pedido,
               SUM({cantidad}) AS unidades,
               SUM({valor}) AS valor
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        {base_where}
        GROUP BY p.codigo
        ORDER BY veces_pedido DESC
        """,
        conn, params=params,
    )

    por_categoria = pd.read_sql(
        f"""
        SELECT COALESCE(p.categoria, '(sin categoría)') AS categoria,
               SUM({cantidad}) AS unidades,
               SUM({valor}) AS valor
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        {base_where}
        GROUP BY p.categoria
        ORDER BY unidades DESC
        """,
        conn, params=params,
    )

    por_mes = pd.read_sql(
        f"""
        SELECT substr(s.fecha_solicitud, 1, 7) AS mes,
               COUNT(DISTINCT s.id) AS solicitudes,
               SUM({cantidad}) AS unidades,
               SUM({valor}) AS valor
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        {base_where}
        GROUP BY mes
        ORDER BY mes
        """,
        conn, params=params,
    )
    conn.close()

    for df in (por_departamento, por_oficina, por_solicitante, por_producto,
               por_categoria, por_mes):
        if not df.empty and "unidades" in df:
            df["unidades"] = df["unidades"].fillna(0).apply(formatear_cantidad)
        if not df.empty and "valor" in df:
            df["valor"] = df["valor"].fillna(0).round(0).astype("int64")

    return {
        "departamento": por_departamento, "oficina": por_oficina,
        "solicitante": por_solicitante, "producto": por_producto,
        "categoria": por_categoria, "mes": por_mes,
    }


def historial_filtrado(desde=None, hasta=None, solicitante=None, area=None, oficina=None,
                       estados=None, producto=None, db_path: str = None) -> pd.DataFrame:
    """
    Una fila por solicitud, con los filtros del buscador del historial.
    desde/hasta son fechas (date o 'YYYY-MM-DD'); el resto son textos que se
    buscan de forma flexible (sin acentos ni mayúsculas, coincidencia parcial).

    producto: filtra las solicitudes que incluyan un insumo determinado. Se
    puede escribir el nombre registrado en el sistema o el código; el filtro se
    aplica en la consulta y no después, para que siga siendo rápido cuando haya
    miles de solicitudes acumuladas.
    """
    db_path = db_path or DB_PATH
    filtro_producto, params = "", []
    if producto and str(producto).strip():
        texto = f"%{str(producto).strip().upper()}%"
        filtro_producto = """
            WHERE EXISTS (
                SELECT 1 FROM solicitud_detalle dd
                JOIN productos pp ON pp.codigo = dd.codigo_producto
                WHERE dd.solicitud_id = s.id
                  AND (UPPER(pp.nombre_estandar) LIKE ? OR pp.codigo LIKE ?)
            )
        """
        params = [texto, texto]

    conn = get_connection(db_path)
    df = pd.read_sql(
        f"""
        SELECT s.folio, s.correlativo, s.fecha_solicitud, s.solicitante,
               s.area_departamento, s.oficina, s.supervisor, s.estado,
               COUNT(d.id) AS n_productos,
               COALESCE(SUM(d.cantidad_solicitada), 0) AS total_unidades
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        {filtro_producto}
        GROUP BY s.id
        ORDER BY s.fecha_solicitud DESC
        """,
        conn, params=params,
    )
    conn.close()
    if df.empty:
        return df

    df["fecha"] = pd.to_datetime(df["fecha_solicitud"], errors="coerce")
    df["total_unidades"] = df["total_unidades"].apply(formatear_cantidad)

    if desde is not None:
        df = df[df["fecha"] >= pd.to_datetime(desde)]
    if hasta is not None:
        # 'hasta' inclusive: se toma hasta el final de ese día
        df = df[df["fecha"] < pd.to_datetime(hasta) + pd.Timedelta(days=1)]

    def _contiene(columna, texto):
        objetivo = normalizar(texto)
        return columna.fillna("").apply(lambda v: objetivo in normalizar(str(v)))

    if solicitante:
        df = df[_contiene(df["solicitante"], solicitante)]
    if area:
        df = df[_contiene(df["area_departamento"], area)]
    if oficina:
        df = df[_contiene(df["oficina"], oficina)]
    if estados:
        df = df[df["estado"].isin(estados)]

    return df


def agrupar_historial(df, agrupacion="mes"):
    """
    Agrega una columna 'periodo' según la agrupación pedida: año, mes,
    semana o día. Devuelve (df con la columna, resumen por período).
    """
    if df.empty:
        return df, pd.DataFrame()

    df = df.copy()
    fechas = pd.to_datetime(df["fecha_solicitud"], errors="coerce")

    if agrupacion == "año":
        df["periodo"] = fechas.dt.strftime("%Y")
    elif agrupacion == "semana":
        # semana ISO, mostrando el lunes de esa semana para que se entienda
        lunes = fechas - pd.to_timedelta(fechas.dt.weekday, unit="D")
        df["periodo"] = ("Semana del " + lunes.dt.strftime("%d/%m/%Y")
                         + " (S" + fechas.dt.isocalendar().week.astype(str) + ")")
    elif agrupacion == "día":
        df["periodo"] = fechas.dt.strftime("%d/%m/%Y")
    else:  # mes
        df["periodo"] = fechas.dt.strftime("%Y-%m")

    resumen = (df.groupby("periodo")
                 .agg(solicitudes=("folio", "count"),
                      productos=("n_productos", "sum"),
                      unidades=("total_unidades", "sum"))
                 .reset_index()
                 .sort_values("periodo", ascending=False))
    return df, resumen


def historial_folios_por_periodo(periodo: str, db_path: str = None) -> pd.DataFrame:
    """
    Una fila POR FOLIO (no por producto): folio, fecha, solicitante, área,
    estado y cuántos productos trae. El detalle de cada folio se pide aparte
    con detalle_folio(), para que el historial no sea una tabla larguísima
    donde una solicitud de 8 productos ocupa 8 filas.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.folio, s.fecha_solicitud, s.solicitante, s.area_departamento,
               s.supervisor, s.estado, COUNT(d.id) AS n_productos,
               COALESCE(SUM(d.cantidad_solicitada), 0) AS total_unidades
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        WHERE s.fecha_solicitud LIKE ?
        GROUP BY s.id
        ORDER BY s.fecha_solicitud DESC
        """,
        conn,
        params=(f"{periodo}%",),
    )
    conn.close()
    if not df.empty:
        df["total_unidades"] = df["total_unidades"].apply(formatear_cantidad)
    return df


def detalle_folio(folio: str, db_path: str = None) -> pd.DataFrame:
    """Los insumos de un folio puntual — se usa al desplegar un folio del historial."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT p.codigo, p.nombre_estandar AS producto, p.unidad_medida,
               d.cantidad_solicitada, d.cantidad_entregada, d.mensaje_sistema
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        WHERE s.folio = ?
        ORDER BY p.nombre_estandar
        """,
        conn,
        params=(folio,),
    )
    conn.close()
    for col in ("cantidad_solicitada", "cantidad_entregada"):
        df[col] = df[col].apply(formatear_cantidad)
    return _normalizar_columna_mensaje(df)


def solicitudes_de(solicitante: str, db_path: str = None) -> pd.DataFrame:
    """Historial de solicitudes de un solicitante específico (para que la
    persona vea el estado de lo que ya pidió, en su propia interfaz)."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT folio, estado, fecha_solicitud, area_departamento
        FROM solicitudes
        WHERE solicitante = ?
        ORDER BY fecha_solicitud DESC
        """,
        conn,
        params=(solicitante,),
    )
    conn.close()
    return df


def solicitudes_de_correo(correo: str, db_path: str = None) -> pd.DataFrame:
    """Igual que solicitudes_de(), pero filtrando por correo registrado en
    vez de nombre escrito a mano — evita que un typo en el nombre esconda
    solicitudes propias."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT folio, estado, fecha_solicitud, area_departamento
        FROM solicitudes
        WHERE correo_solicitante = ?
        ORDER BY fecha_solicitud DESC
        """,
        conn,
        params=(correo.strip().lower(),),
    )
    conn.close()
    return df


def actualizar_valores_catalogo(db_path: str = None):
    """
    Recalcula valor_saldo y valor_unitario de cada producto a partir de
    catalogo_real.py SIN tocar saldo, stock_critico, ni nada más. Existe
    porque una base creada antes de que se agregara la valorización (ver
    migraciones en init_db) queda con valor_saldo/valor_unitario en 0 —
    cargar_catalogo() no se vuelve a correr sobre una base que ya existía, así
    que esos valores nunca se cargaban. Es seguro llamarla siempre al arrancar
    la app: es idempotente y no descuadra el stock ya movido.

    El valor_unitario se deriva del corte original (valor_saldo / saldo del
    catálogo, NO del saldo ya movido en la base), que es justamente lo que
    exige la regla: el unitario se fija en el escaneo y no se recalcula con el
    stock consumido.
    """
    db_path = db_path or DB_PATH
    from catalogo_real import PRODUCTOS

    conn = get_connection(db_path)
    for codigo, _nombre, _unidad, saldo_original, _stock_critico, valor_saldo in PRODUCTOS:
        unitario = valor_unitario_desde(valor_saldo, saldo_original)
        # valor_saldo se deja como valor_unitario * saldo ACTUAL (no el total
        # original): así queda coherente con el stock real: en una base recién
        # cargada el saldo es el original y da el total del corte; en una que
        # ya movió stock, refleja lo que queda. El unitario sí sale del corte
        # original (valor / saldo_original), fijo.
        conn.execute(
            "UPDATE productos SET valor_unitario = ?, valor_saldo = ? * saldo WHERE codigo = ?",
            (unitario, unitario, codigo),
        )
    conn.commit()
    conn.close()
    invalidar_cache_lecturas()


def guardar_config(clave, valor, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute(
        "INSERT INTO configuracion (clave, valor) VALUES (?, ?) "
        "ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor",
        (clave, valor),
    )
    conn.commit()
    conn.close()


def leer_config(clave, por_defecto="", db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    fila = conn.execute("SELECT valor FROM configuracion WHERE clave=?", (clave,)).fetchone()
    conn.close()
    return fila[0] if fila and fila[0] else por_defecto


def nombre_encargado(db_path: str = None) -> str:
    """Nombre y apellido de quien procesa las solicitudes (va en el comprobante)."""
    db_path = db_path or DB_PATH
    return leer_config("nombre_encargado", USUARIO_BODEGA_POR_DEFECTO, db_path)


def texto_info_adicional(cabecera) -> str:
    """
    Texto por defecto del bloque INFORMACIÓN ADICIONAL, con el mismo formato
    que usa hoy el comprobante en papel. El encargado lo edita en pantalla
    antes de descargar: el espacio del medio ya no se imprime en blanco para
    rellenar a mano.
    """
    area = (cabecera.get("area_departamento") or "").strip().upper()
    oficina = (cabecera.get("oficina") or "").strip().upper()
    fecha = cabecera.get("fecha_solicitud") or ""
    try:
        fecha = datetime.strptime(fecha[:10], "%Y-%m-%d").strftime("%d/%m/%y")
    except (ValueError, TypeError):
        pass
    partes = ", ".join(p for p in (area, oficina) if p)
    return (f"{partes}, ENTREGA DE INSUMOS DE OFICINA "
            f"SEGÚN ORDEN ADJUNTA CON FECHA {fecha}")


def guardar_depto_origen(folio, texto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET depto_origen=? WHERE folio=?", (texto, folio))
    conn.commit()
    conn.close()


def guardar_info_adicional(folio, texto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET info_adicional=? WHERE folio=?", (texto, folio))
    conn.commit()
    conn.close()


def guardar_memo(folio, texto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET memo=? WHERE folio=?", (texto, folio))
    conn.commit()
    conn.close()


def guardar_tipo_movimiento(folio, texto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET tipo_movimiento=? WHERE folio=?", (texto, folio))
    conn.commit()
    conn.close()


def guardar_destino(folio, texto, db_path: str = None):
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    conn.execute("UPDATE solicitudes SET destino=? WHERE folio=?", (texto, folio))
    conn.commit()
    conn.close()


def datos_para_impresion(folio, db_path: str = None):
    """
    Devuelve (cabecera: dict, items: list[dict]) con todo lo necesario para
    llenar el formulario físico. Es la fuente única que usa formato_impresion.py.
    """
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    cab = conn.execute(
        "SELECT folio, correlativo, fecha_solicitud, solicitante, supervisor, "
        "       area_departamento, estado, correo_solicitante, correo_supervisor, "
        "       oficina, usuario_operacion, info_adicional, depto_origen, "
        "       memo, tipo_movimiento, destino "
        "FROM solicitudes WHERE folio = ?",
        (folio,),
    ).fetchone()
    if cab is None:
        conn.close()
        return None, []
    filas = conn.execute(
        """
        SELECT p.codigo, p.nombre_estandar, p.unidad_medida,
               d.cantidad_solicitada, d.cantidad_entregada
        FROM solicitud_detalle d
        JOIN productos p ON p.codigo = d.codigo_producto
        JOIN solicitudes s ON s.id = d.solicitud_id
        WHERE s.folio = ?
        ORDER BY p.nombre_estandar
        """,
        (folio,),
    ).fetchall()
    conn.close()

    cabecera = {
        "folio": cab[0], "correlativo": cab[1], "fecha_solicitud": cab[2],
        "solicitante": cab[3], "supervisor": cab[4], "area_departamento": cab[5],
        "estado": cab[6], "correo_solicitante": cab[7], "correo_supervisor": cab[8],
        "oficina": cab[9] or "", "usuario_operacion": cab[10] or nombre_encargado(db_path),
    }
    cabecera["info_adicional"] = cab[11] or texto_info_adicional(cabecera)
    # El nombre formal de la dirección de origen suele diferir del área que
    # escribe el solicitante ("FINANZAS" vs "DIRECCIÓN DE ADMINISTRACIÓN Y
    # FINANZAS"): por eso es editable por el encargado antes de imprimir.
    cabecera["depto_origen"] = cab[12] or cabecera["area_departamento"]
    # Memo, tipo de movimiento y destino: vienen con el valor de siempre por
    # defecto (formato_impresion.py lo rellena si viene vacío), pero quedan
    # editables antes de imprimir el comprobante, igual que depto_origen.
    cabecera["memo"] = cab[13] or ""
    cabecera["tipo_movimiento"] = cab[14] or ""
    cabecera["destino"] = cab[15] or ""
    items = [
        {
            "codigo": f[0], "producto": f[1], "unidad": f[2],
            "cantidad_solicitada": formatear_cantidad(f[3]),
            "cantidad_entregada": formatear_cantidad(f[4]) if f[4] is not None else None,
        }
        for f in filas
    ]
    return cabecera, items


def listar_alertas(tipo=None, db_path: str = None) -> pd.DataFrame:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    if tipo:
        df = pd.read_sql("SELECT * FROM alertas WHERE tipo=? ORDER BY id DESC", conn, params=(tipo,))
    else:
        df = pd.read_sql("SELECT * FROM alertas ORDER BY id DESC", conn)
    conn.close()
    return df


# ----------------------------------------------- sincronización hacia SMC
#
# Estado real (según lo conversado): hoy no está confirmado si SMC ofrece
# alguna vía de integración (import de archivo, ODBC, API). Estas funciones
# NO inventan una conexión que no existe. Lo que hacen es dejar, cada vez
# que se corren, un archivo de intercambio con los movimientos CERRADOS que
# aún no se han marcado como sincronizados — listo para: (a) que alguien lo
# importe a mano en SMC si algún día se confirma que existe esa opción, o
# (b) servir igual como respaldo/auditoría si nunca se conecta. No se marca
# nada como "enviado a SMC" automáticamente; eso requeriría confirmar que el
# archivo efectivamente entró a SMC, lo cual hoy nadie puede verificar desde
# este sistema.

# ============================================================================
#  IMPORTACIÓN DE SALDOS DESDE SMC  (dirección SMC -> web)
# ============================================================================
#
# REGLA DE ORO DE LA SINCRONIZACIÓN
# --------------------------------
# SMC es el sistema de registro oficial del stock; esta web NO lo es.
#
#   * Hacia SMC este sistema envía únicamente MOVIMIENTOS ("salieron 2
#     unidades del código X con el folio Y"), nunca un saldo absoluto. Por eso
#     no puede sobrescribir el inventario real: si bodega recibe 300 rollos de
#     confort y eso se registra en SMC, ese ingreso queda intacto, y el
#     movimiento de salida de esta web simplemente se resta encima.
#     Un saldo absoluto sí podría pisar el ingreso; un movimiento no.
#
#   * Desde SMC este sistema RECIBE el saldo y lo toma como verdad,
#     reemplazando su propia estimación. Es decir, ante cualquier diferencia
#     gana SMC, nunca la web.
#
# Entre una importación y la siguiente, el saldo que muestra la web es una
# ESTIMACIÓN: el saldo del último corte menos lo que ella misma entregó. No
# ve las compras, devoluciones ni ajustes hechos directamente en SMC, y por
# eso siempre se muestra acompañado de su fecha de corte.

def importar_saldos_smc(ruta_archivo, db_path: str = None, fecha_corte=None):
    """
    Actualiza los saldos desde un export de SMC (Excel/CSV del "Listado de
    Saldos Artículos"). Debe traer al menos las columnas 'codigo' y 'saldo';
    opcionalmente 'stock_critico' y 'valor_saldo' (el 'SALDO VAL.' del
    listado — si viene, se actualiza junto con el saldo; si no viene, el
    valor de saldo no se toca y queda como el del último corte conocido).

    El saldo importado REEMPLAZA la estimación local: así, cualquier ingreso
    registrado en SMC (compras, devoluciones, ajustes de inventario) corrige
    automáticamente a la web, en vez de quedar invisible.

    Devuelve (resumen: DataFrame de diferencias, n_actualizados). El resumen
    muestra en qué productos la estimación local se había desviado del saldo
    real y por cuánto — sirve para detectar movimientos hechos fuera de este
    sistema.
    """
    db_path = db_path or DB_PATH
    def _es_csv(r):
        return str(getattr(r, "name", r)).lower().endswith((".csv", ".txt"))

    def _leer(header):
        if _es_csv(ruta_archivo):
            return pd.read_csv(ruta_archivo, dtype=str, header=header)
        return pd.read_excel(ruta_archivo, dtype=str, header=header)

    crudo = _leer(None)
    fila_encabezado = None
    for i in range(min(25, len(crudo))):
        valores = [str(v).strip().lower() for v in crudo.iloc[i].tolist()]
        if "codigo" in valores and "saldo" in valores:
            fila_encabezado = i
            break
    if fila_encabezado is None:
        raise ValueError(
            "No se encontró una fila de títulos con las columnas 'codigo' y 'saldo'. "
            "Revise el archivo exportado desde SMC."
        )

    if hasattr(ruta_archivo, "seek"):
        ruta_archivo.seek(0)
    df = _leer(fila_encabezado)
    df.columns = [str(c).strip().lower() for c in df.columns]

    fecha = fecha_corte or ahora_chile().strftime("%Y-%m-%d %H:%M")
    conn = get_connection(db_path)
    diferencias, actualizados = [], 0

    for _, fila in df.iterrows():
        codigo = fila.get("codigo")
        codigo = codigo.strip() if isinstance(codigo, str) else ""
        if not codigo:
            continue

        def _num(valor):
            if not isinstance(valor, str):
                return None
            valor = valor.strip().replace(".", "").replace(",", ".")
            try:
                return float(valor)
            except ValueError:
                return None

        saldo_nuevo = _num(fila.get("saldo"))
        if saldo_nuevo is None:
            continue
        critico_nuevo = _num(fila.get("stock_critico"))
        valor_nuevo = _num(fila.get("valor_saldo"))
        if valor_nuevo is None:
            valor_nuevo = _num(fila.get("saldo_val"))

        actual = conn.execute(
            "SELECT saldo, nombre_estandar FROM productos WHERE codigo = ?", (codigo,)
        ).fetchone()
        if actual is None:
            diferencias.append({
                "codigo": codigo, "producto": "(no está en el catálogo local)",
                "saldo_estimado": None, "saldo_smc": formatear_cantidad(saldo_nuevo),
                "diferencia": None,
                "observacion": "Producto nuevo en SMC: hay que agregarlo al catálogo.",
            })
            continue

        saldo_estimado, nombre = actual
        delta = formatear_cantidad(saldo_nuevo) - formatear_cantidad(saldo_estimado)
        if delta != 0:
            diferencias.append({
                "codigo": codigo, "producto": nombre,
                "saldo_estimado": formatear_cantidad(saldo_estimado),
                "saldo_smc": formatear_cantidad(saldo_nuevo),
                "diferencia": delta,
                "observacion": ("Ingreso o ajuste registrado en SMC" if delta > 0
                                else "Salida registrada fuera de este sistema"),
            })

        campos = ["saldo=?", "saldo_importado=?", "fecha_corte=?"]
        valores = [saldo_nuevo, saldo_nuevo, fecha]
        if critico_nuevo is not None:
            campos.append("stock_critico=?")
            valores.append(critico_nuevo)
        if valor_nuevo is not None:
            # El archivo trae el valor total (SALDO VAL.): es un corte nuevo
            # CON pesos, así que este es el "momento del escaneo" donde SÍ se
            # refija el valor unitario = total / cantidad de este corte.
            campos.append("valor_saldo=?")
            valores.append(valor_nuevo)
            campos.append("valor_unitario=?")
            valores.append(valor_unitario_desde(valor_nuevo, saldo_nuevo))
        else:
            # Sin valor en el archivo: se mantiene el unitario del corte previo
            # y se reajusta el total = unitario * nuevo_saldo (invariante).
            campos.append("valor_saldo = valor_unitario * ?")
            valores.append(saldo_nuevo)
        valores.append(codigo)
        conn.execute(f"UPDATE productos SET {', '.join(campos)} WHERE codigo=?", valores)
        actualizados += 1

    conn.commit()
    conn.close()
    guardar_config("ultima_importacion_smc", fecha, db_path)
    invalidar_cache_lecturas()  # cambió el saldo/valor de productos
    return pd.DataFrame(diferencias), actualizados


# ============================================================================
#  LECTURA POR OCR DE UN PDF ESCANEADO (papel -> escaneo -> saldos)
# ============================================================================
#
# SMC solo permite imprimir el "Listado de Saldos Artículos" en papel; el
# encargado lo escanea y sube el PDF resultante (sin capa de texto: es una
# foto de cada hoja). Un error de OCR en un saldo es peligroso porque se ve
# igual a un ajuste legítimo hecho en SMC, así que este módulo NUNCA aplica
# nada solo: arma una tabla código/saldo para que el encargado la revise y
# corrija en pantalla, y solo después aplicarla con aplicar_saldos_revisados().
#
# Probado contra un escaneo real de 10 páginas / 290 productos: ~98% de las
# filas salen exactas de punto a punto; el resto (código difícil de leer,
# saldo no reconocido, o un cambio de magnitud/signo inusual respecto al
# saldo actual) queda marcado con revisar=True para que salte a la vista.

_RE_CODIGO_OCR = re.compile(r"^\d{5,9}\.?$")
_RE_NUM_OCR = re.compile(r"[^\d,.\-]*(-?[\d.]+,\d{2})$")


def _num_chile(texto):
    """Convierte '1.234,50' o '-4,00' (formato del listado SMC) a float."""
    if texto is None:
        return None
    texto = str(texto).strip()
    if not texto:
        return None
    negativo = texto.startswith("-")
    texto = texto.lstrip("-").replace(".", "").replace(",", ".")
    try:
        valor = float(texto)
    except ValueError:
        return None
    return -valor if negativo else valor


def _configurar_tesseract():
    """
    En el despliegue (Streamlit Cloud) tesseract se instala vía packages.txt
    y queda en el PATH, así que pytesseract lo encuentra solo. En Windows
    (desarrollo local) no hay una convención de PATH para esto, así que se
    prueba además la ruta típica del instalador de UB-Mannheim.
    """
    import shutil
    import pytesseract
    if shutil.which("tesseract"):
        return
    ruta_windows = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(ruta_windows):
        pytesseract.pytesseract.tesseract_cmd = ruta_windows


def _mejor_header_ocr(palabras, objetivo, y_min=0, y_max=450, umbral=55):
    """Palabra reconocida cuyo texto más se parece a `objetivo` (encabezado
    de columna), dentro de una franja vertical — para tolerar errores de OCR
    en el propio encabezado (ej. 'CODIGO' leído como 'CODED')."""
    from rapidfuzz import fuzz
    mejor, mejor_score = None, 0
    for p in palabras:
        if not (y_min <= p["y"] < y_max):
            continue
        score = fuzz.ratio(p["texto"].strip(".,:").upper(), objetivo)
        if score > mejor_score:
            mejor, mejor_score = p, score
    return mejor if mejor_score >= umbral else None


def _columnas_saldo_ocr(palabras):
    """
    Ubica los límites horizontales de la columna SALDO (la cantidad real,
    no 'SALDO VAL.' en pesos) a partir de los encabezados CODIGO y UNIDAD,
    que son inequívocos. La búsqueda de la palabra 'SALDO' se acota a esa
    misma fila para no confundirla con 'SALDOS' del título del listado
    ("LISTADO DE SALDOS ARTICULOS"), que también calza por aproximación.
    Devuelve (límite_izquierdo, límite_derecho) en píxeles, o None.
    """
    from rapidfuzz import fuzz
    h_codigo = _mejor_header_ocr(palabras, "CODIGO")
    h_unidad = _mejor_header_ocr(palabras, "UNIDAD")
    if not (h_codigo and h_unidad):
        return None
    y_fila = (h_codigo["y"] + h_unidad["y"]) / 2
    cand_saldo = [p for p in palabras if abs(p["y"] - y_fila) <= 20
                  and fuzz.ratio(p["texto"].strip(".,:").upper(), "SALDO") >= 70]
    cand_saldo.sort(key=lambda p: p["x"])
    if len(cand_saldo) < 2:
        return None
    x1, x2 = cand_saldo[0]["x"], cand_saldo[1]["x"]
    return (x1 + x2) / 2, (x2 + h_unidad["x"]) / 2


def extraer_saldos_pdf_escaneado(archivo_pdf, db_path: str = None, idioma: str = "eng") -> pd.DataFrame:
    """
    Lee un PDF escaneado del "Listado de Saldos Artículos" de SMC con OCR y
    arma una tabla código/saldo lista para revisar en pantalla — NO aplica
    nada al inventario; para eso ver aplicar_saldos_revisados().

    archivo_pdf: ruta en disco, o un objeto tipo archivo (ej. lo que entrega
    st.file_uploader), con el contenido del PDF.

    Cada fila queda marcada con 'revisar'=True cuando: el código leído no
    calza exacto con el catálogo, el saldo no se pudo leer, o el cambio
    respecto al saldo actual del sistema es grande o cambia de signo — son
    justo los casos donde un error de OCR pasaría desapercibido si se
    aplicara solo.
    """
    db_path = db_path or DB_PATH
    import difflib

    import fitz
    import pytesseract
    from PIL import Image
    from pytesseract import Output

    _configurar_tesseract()

    conn = get_connection(db_path)
    catalogo = conn.execute("SELECT codigo, nombre_estandar, saldo FROM productos").fetchall()
    conn.close()
    mapa_catalogo = {c: (n, s) for c, n, s in catalogo}
    lista_codigos = list(mapa_catalogo.keys())

    if hasattr(archivo_pdf, "read"):
        doc = fitz.open(stream=archivo_pdf.read(), filetype="pdf")
    else:
        doc = fitz.open(archivo_pdf)

    filas = []
    limites_previos = None

    for i, page in enumerate(doc):
        pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        data = pytesseract.image_to_data(img, lang=idioma, output_type=Output.DICT)
        palabras = [
            {"texto": data["text"][j].strip(), "x": data["left"][j], "y": data["top"][j],
             "conf": int(data["conf"][j])}
            for j in range(len(data["text"])) if data["text"][j].strip()
        ]

        cols = _columnas_saldo_ocr(palabras)
        if cols:
            limites_previos = cols
        elif limites_previos:
            cols = limites_previos
        else:
            continue
        lim_izq, lim_der = cols

        filas_codigo = [p for p in palabras
                         if _RE_CODIGO_OCR.match(p["texto"]) and p["x"] < 300 and p["y"] > 360]
        for fc in filas_codigo:
            y0, y1 = fc["y"] - 15, fc["y"] + 25
            candidatos = []
            for p in palabras:
                if not (lim_izq <= p["x"] < lim_der and y0 <= p["y"] <= y1):
                    continue
                m = _RE_NUM_OCR.search(p["texto"])
                if m:
                    candidatos.append(m.group(1))
            saldo_leido = _num_chile(candidatos[0]) if candidatos else None

            codigo_leido = fc["texto"].rstrip(".")
            coincide = codigo_leido in mapa_catalogo
            sugerido = None
            if not coincide:
                cercanos = difflib.get_close_matches(codigo_leido, lista_codigos, n=1, cutoff=0.75)
                sugerido = cercanos[0] if cercanos else None

            codigo_resuelto = codigo_leido if coincide else sugerido
            nombre = mapa_catalogo[codigo_resuelto][0] if codigo_resuelto else None
            saldo_actual = mapa_catalogo[codigo_resuelto][1] if codigo_resuelto else None

            cambio_grande = False
            if saldo_actual is not None and saldo_leido is not None:
                cambio_signo = (saldo_actual >= 0) != (saldo_leido >= 0)
                cambio_grande = cambio_signo or (
                    abs(saldo_leido - saldo_actual) > max(5, abs(saldo_actual) * 0.3)
                )

            filas.append({
                "pagina": i + 1,
                "aplicar": True,
                "codigo": codigo_resuelto or codigo_leido,
                "codigo_leido": codigo_leido,
                "producto": nombre or "(código no identificado)",
                "saldo_leido": saldo_leido,
                "saldo_actual_sistema": saldo_actual,
                "revisar": (not coincide) or (saldo_leido is None) or cambio_grande,
            })

    df = pd.DataFrame(filas)
    if df.empty:
        return df

    # si dos filas terminan resolviendo al mismo código (típicamente porque
    # una lectura mala "chocó" por aproximación con un código real que
    # también aparece correctamente en otra fila), ambas se marcan para
    # revisar: el conflicto en sí ya es una señal de que algo no calza.
    duplicados = df["codigo"].duplicated(keep=False)
    df.loc[duplicados, "revisar"] = True

    return df.sort_values(["revisar", "pagina"], ascending=[False, True]).reset_index(drop=True)


def aplicar_saldos_revisados(df: pd.DataFrame, db_path: str = None, fecha_corte=None):
    """
    Aplica al inventario real una tabla código/saldo ya revisada por el
    encargado en pantalla (columnas 'codigo' y 'saldo_leido' — el resultado,
    posiblemente editado, de extraer_saldos_pdf_escaneado). Misma lógica de
    reemplazo que importar_saldos_smc: el saldo leído gana sobre la
    estimación local. Si la tabla trae una columna 'aplicar', las filas con
    aplicar=False se descartan (para excluir una fila puntual sin tener que
    borrarla de la tabla).

    Devuelve (diferencias: DataFrame, n_actualizados).
    """
    db_path = db_path or DB_PATH
    if "aplicar" in df.columns:
        df = df[df["aplicar"] != False]  # noqa: E712 (comparación explícita por robustez con NA)

    fecha = fecha_corte or ahora_chile().strftime("%Y-%m-%d %H:%M")
    conn = get_connection(db_path)
    diferencias, actualizados = [], 0

    for _, fila in df.iterrows():
        codigo = str(fila.get("codigo") or "").strip()
        saldo_nuevo = fila.get("saldo_leido")
        if not codigo or saldo_nuevo is None or (isinstance(saldo_nuevo, float) and pd.isna(saldo_nuevo)):
            continue
        try:
            saldo_nuevo = float(saldo_nuevo)
        except (TypeError, ValueError):
            continue

        actual = conn.execute(
            "SELECT saldo, nombre_estandar FROM productos WHERE codigo = ?", (codigo,)
        ).fetchone()
        if actual is None:
            diferencias.append({
                "codigo": codigo, "producto": "(no está en el catálogo local)",
                "saldo_estimado": None, "saldo_smc": formatear_cantidad(saldo_nuevo),
                "diferencia": None,
                "observacion": "Código no existe en el catálogo local: revisar antes de aplicar.",
            })
            continue

        saldo_estimado, nombre = actual
        delta = formatear_cantidad(saldo_nuevo) - formatear_cantidad(saldo_estimado)
        if delta != 0:
            diferencias.append({
                "codigo": codigo, "producto": nombre,
                "saldo_estimado": formatear_cantidad(saldo_estimado),
                "saldo_smc": formatear_cantidad(saldo_nuevo),
                "diferencia": delta,
                "observacion": ("Ingreso o ajuste registrado en SMC" if delta > 0
                                else "Salida registrada fuera de este sistema"),
            })
        # El escaneo OCR trae solo la cantidad, no el valor en pesos, así que
        # NO se recalcula el valor unitario (queda el del corte, como debe).
        # Pero sí se reajusta valor_saldo = valor_unitario * nuevo_saldo para
        # mantener el invariante: el valor en pesos del inventario acompaña a
        # la cantidad reescaneada, con el unitario fijo.
        conn.execute(
            "UPDATE productos SET saldo=?, saldo_importado=?, fecha_corte=?, "
            "valor_saldo = valor_unitario * ? WHERE codigo=?",
            (saldo_nuevo, saldo_nuevo, fecha, saldo_nuevo, codigo),
        )
        actualizados += 1

    conn.commit()
    conn.close()
    guardar_config("ultima_importacion_smc", fecha, db_path)
    invalidar_cache_lecturas()  # cambió el saldo/valor de productos
    return pd.DataFrame(diferencias), actualizados


def horas_desde_ultima_importacion(db_path: str = None):
    """
    Horas transcurridas desde la última importación de saldos desde SMC.
    Devuelve None si nunca se importó.
    """
    db_path = db_path or DB_PATH
    valor = leer_config("ultima_importacion_smc", "", db_path)
    if not valor:
        return None
    try:
        ultima = datetime.strptime(str(valor)[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        return None
    return (ahora_chile() - ultima).total_seconds() / 3600


def fecha_ultimo_corte(db_path: str = None):
    """Fecha del último saldo importado desde SMC, o None si nunca se importó."""
    db_path = db_path or DB_PATH
    valor = leer_config("ultima_importacion_smc", "", db_path)
    return valor or None


HORA_INICIO_JORNADA = 8  # 8:00, junto con HORA_CIERRE_JORNADA (19:00)


def inicio_semana_actual(ahora=None):
    """
    Lunes de la semana vigente a las 8:00 (hora de Chile) — el "comienzo de
    la jornada semanal". Si hoy es lunes antes de las 8:00, todavía cuenta
    como parte de la semana anterior.
    """
    ahora = ahora or ahora_chile()
    lunes = ahora - timedelta(days=ahora.weekday())  # weekday(): lunes=0
    corte = lunes.replace(hour=HORA_INICIO_JORNADA, minute=0, second=0, microsecond=0)
    if ahora < corte:
        corte -= timedelta(days=7)
    return corte


def actualizacion_semanal_pendiente(db_path: str = None) -> bool:
    """
    True si todavía no se ha importado un saldo desde SMC en la semana
    vigente (desde el lunes 8:00). Reemplaza la alerta anterior de "cada 6
    horas": como ahora la importación depende de escanear el listado en
    papel, tiene más sentido pedirlo una vez por semana, calzado con el
    inicio de la jornada semanal, que seguir insistiendo cada pocas horas.
    """
    db_path = db_path or DB_PATH
    valor = leer_config("ultima_importacion_smc", "", db_path)
    if not valor:
        return True
    try:
        ultima = datetime.strptime(str(valor)[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        return True
    return ultima < inicio_semana_actual()


RUTA_EXCEL_HISTORIAL = os.path.join("exportes", "historial_bodega.xlsx")
HORAS_ENTRE_EXPORTES_EXCEL = 5  # cadencia pedida para refrescar el Excel


def horas_desde_ultimo_export_excel(db_path: str = None):
    """Horas desde la última vez que se generó el Excel del historial, o None
    si nunca se ha generado."""
    db_path = db_path or DB_PATH
    valor = leer_config("ultima_generacion_excel_historial", "", db_path)
    if not valor:
        return None
    try:
        ultima = datetime.strptime(str(valor)[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        return None
    return (ahora_chile() - ultima).total_seconds() / 3600


def exportar_historial_excel(ruta: str = None, carpeta_pdf: str = "formularios",
                              db_path: str = None) -> str:
    """
    Genera (sobrescribiendo si ya existe) un Excel a partir de los
    COMPROBANTES ENTREGADOS, no de las solicitudes en bruto: solo incluye
    solicitudes cerradas (las únicas que tienen comprobante) y usa
    cantidad_entregada como la cantidad real — el número que el encargado
    dejó después de ajustar o rechazar algún insumo en bodega, que puede
    diferir de lo que el solicitante pidió originalmente. Es más preciso
    porque refleja lo que efectivamente salió de bodega, no lo pedido.

    Una hoja "Detalle" con una fila por producto entregado, dejada como Tabla
    de Excel con autofiltro para que armar una tabla dinámica sea directo
    (Insertar > Tabla dinámica, usando esa tabla como origen), más hojas ya
    resumidas por producto, solicitante, departamento y mes.

    No se adjuntan los PDF de solicitud/comprobante -inflarían mucho el
    archivo-, solo el nombre del archivo correspondiente dentro de la carpeta
    de formularios, para poder ubicarlo aparte si se necesita.
    """
    db_path = db_path or DB_PATH
    ruta = ruta or RUTA_EXCEL_HISTORIAL
    carpeta_destino = os.path.dirname(ruta)
    if carpeta_destino:
        os.makedirs(carpeta_destino, exist_ok=True)

    conn = get_connection(db_path)
    detalle = pd.read_sql(
        """
        SELECT s.correlativo AS "N", s.folio, s.fecha_solicitud AS fecha,
               s.solicitante, s.area_departamento AS departamento, s.oficina,
               s.supervisor, s.usuario_operacion AS entregado_por,
               p.codigo, p.nombre_estandar AS producto, p.categoria,
               p.unidad_medida AS unidad,
               d.cantidad_solicitada AS solicitado, d.cantidad_entregada AS entregado,
               p.valor_unitario AS valor_unitario, d.valor_movimiento AS valor_entregado,
               d.mensaje_sistema AS observacion
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        WHERE s.estado = 'cerrada'
        ORDER BY s.correlativo ASC
        """,
        conn,
    )
    conn.close()

    # Valor en pesos de cada línea entregada. valor_movimiento se fija al
    # cerrar; para comprobantes cerrados antes de existir esta columna (NULL),
    # se estima con el valor unitario actual por cantidad, para no dejar la
    # celda vacía en el histórico.
    if not detalle.empty:
        est = (detalle["valor_unitario"].fillna(0) * detalle["entregado"].fillna(0)).round(0)
        detalle["valor_entregado"] = detalle["valor_entregado"].fillna(est).round(0)
        detalle["valor_unitario"] = detalle["valor_unitario"].fillna(0).round(0)

    def _nombre_si_existe(prefijo, correlativo):
        nombre = f"{prefijo}_{correlativo}.pdf"
        return nombre if os.path.exists(os.path.join(carpeta_pdf, nombre)) else ""

    if not detalle.empty:
        detalle["archivo_solicitud"] = detalle["N"].apply(lambda c: _nombre_si_existe("solicitud", c))
        detalle["archivo_comprobante"] = detalle["N"].apply(lambda c: _nombre_si_existe("comprobante", c))

    resumenes = estadisticas_consumo(solo_cerradas=True, db_path=db_path)

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        detalle.to_excel(writer, sheet_name="Detalle", index=False)
        resumenes["producto"].to_excel(writer, sheet_name="Por producto", index=False)
        resumenes["solicitante"].to_excel(writer, sheet_name="Por solicitante", index=False)
        resumenes["departamento"].to_excel(writer, sheet_name="Por departamento", index=False)
        resumenes["mes"].to_excel(writer, sheet_name="Por mes", index=False)

        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Una hoja por MES CALENDARIO (Enero..Diciembre), además de "Detalle"
        # (todo junto) y "Por mes" (resumen). Cada hoja junta TODOS los pedidos
        # de ese mes sin importar el año (todos los eneros en "Enero", etc.),
        # ordenados por fecha ascendente. El orden es estable, así las líneas de
        # un mismo pedido (que comparten fecha) quedan contiguas y siguen
        # funcionando el coloreado por pedido y el ditto. Las líneas sin fecha
        # reconocible, si las hubiera, van a una hoja "Sin fecha".
        MESES_ES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                    9: "Septiembre", 10: "Octubre", 11: "Noviembre",
                    12: "Diciembre"}
        hojas_mes = []  # [(nombre_hoja, df_mes, nombre_tabla)]
        if not detalle.empty:
            fechas = pd.to_datetime(detalle["fecha"], errors="coerce")
            mes_num = fechas.dt.month
            for numero in sorted({int(m) for m in mes_num.dropna().unique()}):
                mascara = mes_num == numero
                orden = fechas[mascara].sort_values(kind="stable").index  # cronológico
                df_mes = detalle.loc[orden]
                hojas_mes.append((MESES_ES[numero], df_mes, f"Tabla_{numero:02d}"))
            sin_fecha = detalle[mes_num.isna()]
            if not sin_fecha.empty:
                hojas_mes.append(("Sin fecha", sin_fecha, "Tabla_sin_fecha"))
        for nombre_hoja, df_mes, _tabla in hojas_mes:
            df_mes.to_excel(writer, sheet_name=nombre_hoja, index=False)

        # Formato de dinero en pesos chilenos (sin decimales, separador de miles).
        FORMATO_CLP = '"$"#,##0'

        def _es_columna_valor(nombre) -> bool:
            return "valor" in str(nombre).strip().lower()

        def _aplicar_formato_pesos(hoja, df):
            """Pone formato CLP a toda columna cuyo encabezado tenga 'valor'
            (valor, valor_unitario, valor_total, valor_entregado, etc.)."""
            for indice, encabezado in enumerate(df.columns, start=1):
                if _es_columna_valor(encabezado):
                    letra = get_column_letter(indice)
                    for fila in range(2, len(df) + 2):  # fila 1 = encabezado
                        hoja[f"{letra}{fila}"].number_format = FORMATO_CLP

        def _autoajustar_anchos(hoja, tope=40):
            """Ancho de cada columna = largo del texto más largo (encabezado o
            dato), con un tope para que una observación larga no la agrande de
            más."""
            for columna in hoja.columns:
                ancho = max((len(str(c.value)) for c in columna if c.value is not None), default=0)
                hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 2, tope)

        # Columnas de cabecera del pedido que se colapsan a "" en las líneas
        # repetidas de un mismo pedido (índices 1-based, calculados por nombre
        # por si cambia el orden de las columnas de la consulta).
        COLS_CABECERA_PEDIDO = ["folio", "fecha", "solicitante", "departamento",
                                 "oficina", "supervisor", "entregado_por"]
        fill_blanco = PatternFill("solid", fgColor="FFFFFFFF")
        fill_celeste = PatternFill("solid", fgColor="FFDCE6F1")

        def _formatear_detalle(hoja, df, nombre_tabla):
            """Deja una hoja de detalle como Tabla de Excel (para dinámicas) con
            colores alternados POR PEDIDO —cada folio comparte color, alternando
            blanco/#FFFFFF y celeste/#DCE6F1— y la cabecera del pedido colapsada
            a "" en las líneas repetidas. Se apaga el rayado por fila de la tabla
            (showRowStripes=False) para no pisar el alternado por pedido. Nota:
            el ditto rompe las tablas dinámicas que agrupen por esas columnas."""
            n_filas, n_cols = df.shape
            if n_filas == 0:
                return
            ultima_col = get_column_letter(n_cols)
            tabla = Table(displayName=nombre_tabla, ref=f"A1:{ultima_col}{n_filas + 1}")
            tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
            hoja.add_table(tabla)
            indices_cabecera = [df.columns.get_loc(c) + 1
                                for c in COLS_CABECERA_PEDIDO if c in df.columns]
            folio_previo = object()  # centinela distinto de cualquier folio real
            usar_celeste = False
            for i, folio_val in enumerate(df["folio"].tolist()):
                es_primera_linea = folio_val != folio_previo
                if es_primera_linea:
                    usar_celeste = not usar_celeste
                    folio_previo = folio_val
                relleno = fill_celeste if usar_celeste else fill_blanco
                for col in range(1, n_cols + 1):
                    hoja.cell(row=i + 2, column=col).fill = relleno  # +2: fila 1 = encabezado
                if not es_primera_linea:  # línea repetida del pedido -> "" (ditto)
                    for col in indices_cabecera:
                        hoja.cell(row=i + 2, column=col).value = '""'

        hojas = {
            "Detalle": detalle,
            "Por producto": resumenes["producto"],
            "Por solicitante": resumenes["solicitante"],
            "Por departamento": resumenes["departamento"],
            "Por mes": resumenes["mes"],
        }
        # Formato CLP y auto-ajuste de anchos en TODAS las hojas (base + mensuales).
        # El auto-ajuste corre antes del ditto, así mide los valores completos.
        for nombre_hoja, df_hoja in list(hojas.items()) + [(n, d) for n, d, _ in hojas_mes]:
            if df_hoja is not None and not df_hoja.empty:
                _aplicar_formato_pesos(writer.sheets[nombre_hoja], df_hoja)
                _autoajustar_anchos(writer.sheets[nombre_hoja])

        # Anchos fijos de "Por producto": A=400 px, B=200 px. Excel mide el ancho
        # en caracteres, no en píxeles: se convierte con (px - 5) / 7 para la
        # fuente por defecto. Va después del auto-ajuste para que estos ganen.
        def _px_a_ancho(px):
            return round((px - 5) / 7.0, 2)
        if not resumenes["producto"].empty:
            hp = writer.sheets["Por producto"]
            hp.column_dimensions["A"].width = _px_a_ancho(400)
            hp.column_dimensions["B"].width = _px_a_ancho(200)

        # Tabla + colores por pedido + ditto: en la hoja Detalle (todo junto) y
        # en cada hoja mensual.
        if not detalle.empty:
            _formatear_detalle(writer.sheets["Detalle"], detalle, "TablaHistorial")
        for nombre_hoja, df_mes, nombre_tabla in hojas_mes:
            _formatear_detalle(writer.sheets[nombre_hoja], df_mes, nombre_tabla)

    guardar_config("ultima_generacion_excel_historial",
                    ahora_chile().strftime("%Y-%m-%d %H:%M"), db_path)
    return ruta


def obtener_pendientes_sync(db_path: str = None) -> pd.DataFrame:
    """Solicitudes cerradas que aún no se han incluido en ningún archivo
    de sincronización hacia SMC."""
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql(
        """
        SELECT s.folio, s.fecha_solicitud, s.solicitante, s.area_departamento,
               p.codigo, p.nombre_estandar AS producto, d.cantidad_entregada
        FROM solicitudes s
        JOIN solicitud_detalle d ON d.solicitud_id = s.id
        JOIN productos p ON p.codigo = d.codigo_producto
        WHERE s.estado = 'cerrada' AND s.sincronizado_smc = 0
        ORDER BY s.fecha_solicitud
        """,
        conn,
    )
    conn.close()
    if not df.empty:
        df["cantidad_entregada"] = df["cantidad_entregada"].apply(formatear_cantidad)
    return df


def generar_archivo_sync_smc(carpeta_salida: str = ".", db_path: str = None):
    """
    Genera un CSV con los movimientos pendientes (formato: código, cantidad
    entregada, folio, fecha) y deja registro en log_sincronizacion_smc.
    Devuelve (ruta_archivo, n_folios) o (None, 0) si no había nada pendiente.

    Pensado para correr cada ~10 min vía un scheduler (Programador de tareas
    de Windows, cron, o un loop simple con la librería `schedule`) — ver
    sincronizar_smc.py.
    """
    db_path = db_path or DB_PATH
    import os as _os

    pendientes = obtener_pendientes_sync(db_path)
    if pendientes.empty:
        return None, 0

    fecha = ahora_chile().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"sync_smc_{fecha}.csv"
    ruta = _os.path.join(carpeta_salida, nombre_archivo)
    pendientes.to_csv(ruta, index=False, encoding="utf-8-sig")

    folios = pendientes["folio"].unique().tolist()
    conn = get_connection(db_path)
    conn.executemany(
        "UPDATE solicitudes SET sincronizado_smc = 1 WHERE folio = ?",
        [(f,) for f in folios],
    )
    conn.execute(
        "INSERT INTO log_sincronizacion_smc (fecha, folios_incluidos, archivo_generado, estado) "
        "VALUES (?, ?, ?, 'exportado_local')",
        (ahora_chile().strftime("%Y-%m-%d %H:%M"), len(folios), nombre_archivo),
    )
    conn.commit()
    conn.close()
    return ruta, len(folios)


def historial_sincronizacion(db_path: str = None) -> pd.DataFrame:
    db_path = db_path or DB_PATH
    conn = get_connection(db_path)
    df = pd.read_sql("SELECT * FROM log_sincronizacion_smc ORDER BY id DESC", conn)
    conn.close()
    return df
